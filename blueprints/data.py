"""Import, export, and completion-sync routes."""

import csv
import json
import io
import logging
import re
from datetime import date
from typing import TypedDict

import openpyxl
from flask import (
    Blueprint,
    Response,
    flash,
    redirect,
    render_template,
    request,
    session,
    url_for,
)
from sqlalchemy import desc

from constants import (
    EDGE_TYPES,
    STATUS_OPTIONS,
    TRUTHY,
    EDGELESS_CATEGORIES,
    UNKNOWN_COLOR,
    XLSX_COL_MAP,
    accepts_set_handle_variants,
    canonicalize_availability,
    canonicalize_category,
    is_edgeless_category,
    normalize_edge_for_category,
    VARIANT_SYNC_SINGLE_VARIANT_CATEGORIES,
)
from extensions import db
from number_utils import parse_positive_whole_number
from blueprints.data_import import (
    _availability_preview_fields,
    _build_item_name_lookup,
    _build_item_sku_lookup,
    _build_notes,
    _build_set_sku_lookup,
    _group_import_rows,
    _match_import_item,
    _merge_note_text,
    _normalize_import_color,
    _normalize_variant_lookup_name,
    _parse_owned_raw,
    _parse_truthy_field,
    _preview_import_color,
    _read_engraving_fields,
    _read_completion_rows,
    _resolve_import_variant_color,
)
from blueprints.data_workflows import (
    _add_import_ownership_quantities,
    _build_completion_missing_csv,
    _build_completion_missing_rows,
    _build_completion_preview,
    _build_import_header_report,
    _find_import_variant,
    _import_row_label,
    _merge_import_ownership,
    _parse_quantity_fields,
    _read_confirm_quantity_field,
    _resolve_completion_gap_people,
    _safe_csv_filename,
)
from helpers import admin_required, db_commit
from models import (
    Item,
    ItemSetMember,
    ItemVariant,
    Ownership,
    ActivityEvent,
    Person,
    Set,
    get_or_create_set,
    normalize_sku_value,
    record_activity,
    reconcile_unknown_variant,
)
from scraping import (
    scrape_item_variant_colors,  # noqa: F401
    scrape_set_variant_options,  # noqa: F401
)
from scraping import scrape_purple_campaign_variants  # noqa: F401
from time_utils import format_container_time

data_bp = Blueprint("data", __name__)
logger = logging.getLogger(__name__)


class SetMemberEntry(TypedDict):
    """Typed row fragment for set membership persistence."""

    sku: str
    quantity: int
    name: str


def _parse_set_members_field(raw_value: str | None) -> list[str]:
    """Split a set_members cell into normalized member SKUs."""
    if not raw_value:
        return []
    members: list[str] = []
    for chunk in re.split(r"[|,\n]+", raw_value):
        sku = normalize_sku_value(chunk)
        if sku:
            members.append(sku)
    return members


def _format_set_members_display(member_skus: list[str]) -> str:
    """Build a compact label for a set member SKU list."""
    if not member_skus:
        return "—"
    counts: dict[str, int] = {}
    order: list[str] = []
    for sku in member_skus:
        if sku not in counts:
            order.append(sku)
            counts[sku] = 0
        counts[sku] += 1
    return " | ".join(
        f"{sku} ×{counts[sku]}" if counts[sku] > 1 else sku for sku in order
    )


def _build_import_skip_detail(
    row_num: int | None,
    reason: str,
    *,
    name: str | None = None,
    sku: str | None = None,
) -> dict:
    """Build a standard skipped-row payload for import confirmations."""
    return {
        "row": row_num,
        "label": _import_row_label(row_num, name, sku),
        "reason": reason,
    }


def _append_import_skip_detail(
    skipped_details: list[dict],
    row_num: int | None,
    reason: str,
    *,
    name: str | None = None,
    sku: str | None = None,
) -> None:
    """Append a standard skipped-row payload to the running detail list."""
    skipped_details.append(
        _build_import_skip_detail(row_num, reason, name=name, sku=sku)
    )


@data_bp.route("/export")
@admin_required
def export_page():
    """Render the export page."""
    suggested_name = f"cutco_collection_{date.today().isoformat()}.csv"
    return render_template("export_page.html", suggested_name=suggested_name)


@data_bp.route("/export/csv")
@admin_required
def export_csv():
    """Export the collection as CSV."""
    rows = (
        db.session.query(Ownership, ItemVariant, Item, Person)
        .join(ItemVariant, Ownership.variant_id == ItemVariant.id)
        .join(Item, ItemVariant.item_id == Item.id)
        .join(Person, Ownership.person_id == Person.id)
        .order_by(Person.name, Item.name, ItemVariant.color, Ownership.copy_type)
        .all()
    )

    csv_buffer = io.StringIO()
    writer = csv.writer(csv_buffer)
    writer.writerow(
        [
            "person",
            "item_name",
            "sku",
            "category",
            "edge_type",
            "color",
            "status",
            "is_sku_unicorn",
            "is_variant_unicorn",
            "is_edge_unicorn",
            "quantity_purchased",
            "quantity_given_away",
            "copy_type",
            "engraving_text",
            "engraving_notes",
            "notes",
        ]
    )
    for ownership, variant, item, person in rows:
        writer.writerow(
            [
                person.name,
                item.name,
                item.sku or "",
                item.category or "",
                item.edge_type,
                variant.color,
                ownership.status,
                "yes" if item.is_unicorn else "no",
                "yes" if variant.is_unicorn else "no",
                "yes" if item.edge_is_unicorn else "no",
                (
                    ownership.quantity_purchased
                    if ownership.quantity_purchased is not None
                    else ""
                ),
                (
                    ownership.quantity_given_away
                    if ownership.quantity_given_away is not None
                    else ""
                ),
                ownership.copy_type,
                ownership.engraving_text or "",
                ownership.engraving_notes or "",
                ownership.notes or "",
            ]
        )
    csv_buffer.seek(0)
    filename = _safe_csv_filename(request.args.get("filename", "cutco_collection.csv"))
    logger.info("CSV export requested: %d rows (%s)", len(rows), filename)
    return Response(
        csv_buffer.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@data_bp.route("/completion-gaps", methods=["GET", "POST"])
@admin_required
def completion_gaps_page():
    """Render the completion gaps page."""
    people = Person.query.order_by(Person.name).all()
    public_catalog_count = Item.query.filter_by(set_only=False, in_catalog=True).count()
    last_person_id = session.get("last_person_id")
    default_person_id = (
        last_person_id
        if any(person.id == last_person_id for person in people)
        else "all"
    )

    if request.method == "GET":
        selected_person_id = str(
            request.args.get("person_id") or default_person_id or "all"
        ).strip()
        selected_people, selected_person_value, selection_error = (
            _resolve_completion_gap_people(selected_person_id, people)
        )
        view_mode = (request.args.get("view") or "").strip().lower()
        if selection_error:
            flash(selection_error, "error")
        missing_rows = None
        missing_rows_csv = None
        if view_mode == "screen" and not selection_error:
            missing_rows = _build_completion_missing_rows(
                [person.name for person in selected_people]
            )
            missing_rows_csv = _build_completion_missing_csv(missing_rows)
        return render_template(
            "completion_gaps.html",
            people=people,
            public_catalog_count=public_catalog_count,
            default_person_id=selected_person_value,
            missing_rows=missing_rows,
            missing_rows_csv=missing_rows_csv,
            view_mode=view_mode,
        )

    selected_person_id = str(request.form.get("person_id") or "all").strip()
    selected_people, selected_person_value, selection_error = (
        _resolve_completion_gap_people(selected_person_id, people)
    )
    if selection_error:
        flash(selection_error, "error")
        return render_template(
            "completion_gaps.html",
            people=people,
            public_catalog_count=public_catalog_count,
            default_person_id=selected_person_value,
            missing_rows=None,
            missing_rows_csv=None,
            view_mode="",
        )

    filename_prefix = (
        "all_collectors"
        if selected_person_value == "all"
        else selected_people[0].name or "collector"
    )
    missing_rows = _build_completion_missing_rows(
        [person.name for person in selected_people]
    )
    csv_text = _build_completion_missing_csv(missing_rows)
    filename = _safe_csv_filename(
        f"cutco_completion_gaps_{filename_prefix}_{date.today().isoformat()}.csv"
    )
    logger.info(
        "Completion gaps export requested: %d rows (%s)", len(missing_rows), filename
    )
    return Response(
        csv_text,
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@data_bp.route("/import/template")
@admin_required
def import_template():
    """Download a starter CSV template for imports."""
    csv_buffer = io.StringIO()
    writer = csv.writer(csv_buffer)
    writer.writerow(
        [
            "name",
            "sku",
            "owned",
            "color",
            "availability",
            "quantity purchased",
            "quantity given away",
            "category",
            "edge",
            "copy_type",
            "engraved",
            "engraving_text",
            "engraving_notes",
            "is_sku_unicorn",
            "is_variant_unicorn",
            "is_edge_unicorn",
            "set_members",
            "price",
        ]
    )
    writer.writerow(
        [
            '2-3/4" Paring Knife',
            "1720",
            "Anthony",
            "Classic Brown",
            "public",
            "1",
            "0",
            "Kitchen Knives",
            "Double-D",
            "plain",
            "",
            "",
            "",
            "no",
            "no",
            "no",
            "",
            "12.50",
        ]
    )
    writer.writerow(
        [
            "Super Shears",
            "2137",
            "yes",
            "Pearl White",
            "non-catalog",
            "",
            "",
            "Kitchen Knives",
            "Straight",
            "engraved",
            "yes",
            "250th Anniversary",
            "Limited edition engraving",
            "no",
            "no",
            "no",
            "",
            "",
        ]
    )
    csv_buffer.seek(0)
    return Response(
        csv_buffer.getvalue(),
        mimetype="text/csv",
        headers={
            "Content-Disposition": "attachment; filename=cutco_import_starter.csv"
        },
    )


@data_bp.route("/import", methods=["GET", "POST"])
@admin_required
def import_page():
    """Render the import page."""
    if request.method == "GET":
        return render_template(
            "import_page.html",
            people=Person.query.order_by(Person.name).all(),
            import_check=None,
        )

    uploaded_file = request.files.get("csvfile")
    if not uploaded_file or not uploaded_file.filename:
        flash("Please choose a file.", "error")
        return render_template(
            "import_page.html",
            people=Person.query.order_by(Person.name).all(),
            import_check=None,
        )

    person_override = request.form.get("person_override", "").strip() or None
    ext = uploaded_file.filename.rsplit(".", 1)[-1].lower()
    logger.info(
        "Import file received: %s (person override: %s)",
        uploaded_file.filename,
        person_override or "none",
    )

    if request.form.get("mode") == "check":
        try:
            header_report = _build_import_header_report(uploaded_file, ext)
            if header_report["ok"]:
                flash("Header check passed.", "success")
            else:
                flash("Header check found required column issues.", "warning")
            return render_template(
                "import_page.html",
                people=Person.query.order_by(Person.name).all(),
                import_check=header_report,
            )
        except Exception as exc:
            logger.error("Import header check failed: %s", exc)
            flash(
                "Could not read headers from this file. Use CSV/XLSX with a header row.",
                "error",
            )
            return render_template(
                "import_page.html",
                people=Person.query.order_by(Person.name).all(),
                import_check=None,
            )

    try:
        if ext == "xlsx":
            wb = openpyxl.load_workbook(
                io.BytesIO(uploaded_file.stream.read()), data_only=True
            )
            ws = wb.active
            if ws is None:
                raise ValueError("Workbook has no active worksheet")
            raw_headers = [
                str(cell.value).strip() if cell.value is not None else ""
                for cell in ws[1]
            ]
            norm_rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if all(cell_value is None for cell_value in row):
                    continue
                norm_rows.append(
                    {
                        raw_headers[col_idx]: (
                            str(cell_value).strip() if cell_value is not None else ""
                        )
                        for col_idx, cell_value in enumerate(row)
                    }
                )
            parsed_rows = []
            for row in norm_rows:
                out_row = {}
                for orig_key, val in row.items():
                    normalized_key = orig_key.strip().lower()
                    if normalized_key in XLSX_COL_MAP:
                        out_row[XLSX_COL_MAP[normalized_key]] = val
                    else:
                        out_row[normalized_key.replace(" ", "_")] = val
                parsed_rows.append(out_row)
        else:
            stream = io.StringIO(uploaded_file.stream.read().decode("utf-8-sig"))
            reader = csv.DictReader(stream)
            parsed_rows = []
            for row in reader:
                out_row = {}
                for orig_key, val in row.items():
                    normalized_key = orig_key.strip().lower()
                    value = val.strip() if val is not None else ""
                    if normalized_key in XLSX_COL_MAP:
                        out_row[XLSX_COL_MAP[normalized_key]] = value
                    else:
                        out_row[normalized_key.replace(" ", "_")] = value
                parsed_rows.append(out_row)

    except Exception as exc:
        logger.error("Import file parse failed: %s", exc)
        flash(
            "Could not parse the uploaded file — check that it is a valid CSV or XLSX.",
            "error",
        )
        return render_template(
            "import_page.html", people=Person.query.order_by(Person.name).all()
        )

    if person_override:
        for row in parsed_rows:
            row["owned_raw"] = row.get("owned_raw", "yes")
            row["_person_override"] = person_override

    existing_items = _build_item_sku_lookup(Item.query.all())
    existing_set_skus = _build_set_sku_lookup(Set.query.all())
    existing_names = _build_item_name_lookup(
        [item for item in Item.query.all() if not normalize_sku_value(item.sku)]
    )
    existing_persons = {person.name.lower(): person for person in Person.query.all()}

    already_in_catalog = []
    sku_name_mismatches = []
    new_items_list = []
    likely_unicorns = []
    set_sku_collisions = []
    set_rows = []
    ownership_entries = []
    conflicts = []
    errors = []

    for row_num, row in enumerate(parsed_rows, start=2):
        name = row.get("name", "").strip()
        sku = normalize_sku_value(row.get("sku", ""))
        color = _normalize_import_color(row.get("color", ""))
        availability_raw = row.get("availability", "").strip()
        availability = canonicalize_availability(availability_raw)
        legacy_non_catalog = _parse_truthy_field(row.get("non_catalog", ""))
        edge_type = row.get("edge_type", "").strip() or "Unknown"
        set_member_skus = _parse_set_members_field(row.get("set_members", ""))
        is_sku_unicorn = (
            row.get("is_sku_unicorn", row.get("item_is_unicorn", "")).strip().lower()
            in TRUTHY
        )
        is_variant_unicorn = row.get("is_variant_unicorn", "").strip().lower() in TRUTHY
        is_edge_unicorn = (
            row.get("is_edge_unicorn", row.get("edge_is_unicorn", "")).strip().lower()
            in TRUTHY
        )
        non_catalog = (
            legacy_non_catalog
            or is_sku_unicorn
            or is_variant_unicorn
            or is_edge_unicorn
        )
        if availability == "public" and non_catalog:
            availability = "non-catalog"
        non_catalog = non_catalog or availability != "public"
        category = canonicalize_category(row.get("category", ""))
        edge_type, is_edge_unicorn = normalize_edge_for_category(
            category,
            edge_type,
            is_edge_unicorn,
        )
        matched_set = existing_set_skus.get(sku) if sku else None
        is_set_row = bool(set_member_skus)
        if is_edgeless_category(category):
            non_catalog = (
                legacy_non_catalog
                or is_sku_unicorn
                or is_variant_unicorn
                or availability != "public"
            )
        note_text, note_errors = _build_notes(row)
        quantity_purchased, quantity_given_away, quantity_errors = (
            _parse_quantity_fields(row)
        )
        note_errors.extend(quantity_errors)
        copy_type, engraving_text, engraving_notes, engraving_signature = (
            _read_engraving_fields(row)
        )
        if note_errors:
            errors.append(
                {
                    "row": row_num,
                    "reason": "; ".join(note_errors),
                    "data": row,
                }
            )
            continue
        notes = note_text or row.get("notes", "").strip() or None
        owned_raw = row.get("owned_raw", row.get("status", "yes"))
        status, person_name = _parse_owned_raw(
            owned_raw, row.get("_person_override") or row.get("person", "")
        )

        if person_override:
            person_name = person_override

        if not name:
            errors.append({"row": row_num, "reason": "Missing name", "data": row})
            continue

        if is_set_row:
            set_rows.append(
                {
                    "name": name,
                    "sku": sku,
                    "row": row_num,
                    "set_members": set_member_skus,
                    "set_members_display": _format_set_members_display(set_member_skus),
                    "member_count": len(set_member_skus),
                    "matches_existing_set_sku": bool(matched_set),
                    "matched_set_name": (
                        matched_set.name if matched_set is not None else None
                    ),
                    "notes": notes,
                }
            )
            continue

        if status not in STATUS_OPTIONS:
            status = "Owned"

        matched_item = _match_import_item(
            existing_items=existing_items,
            existing_names=existing_names,
            sku=sku,
            name=name,
        )

        matches_set_sku = bool(sku and matched_set and not matched_item)

        item_category_for_color = (
            matched_item.category if matched_item else category
        ) or ""
        target_color = _resolve_import_variant_color(
            name, item_category_for_color, color
        )
        is_cookware = item_category_for_color in VARIANT_SYNC_SINGLE_VARIANT_CATEGORIES
        existing_variant = None
        if matched_item:
            existing_variant = next(
                (
                    variant
                    for variant in matched_item.variants
                    if variant.color.lower() == target_color.lower()
                ),
                None,
            )

        if matched_item:
            already_in_catalog.append(
                {
                    "item": matched_item,
                    "row": row,
                    "row_num": row_num,
                    "color": target_color,
                    "display_color": _preview_import_color(target_color, is_cookware),
                    "non_catalog": non_catalog,
                    "availability": availability,
                    "availability_label": _availability_preview_fields(availability)[0],
                    "availability_badge_class": _availability_preview_fields(
                        availability
                    )[1],
                    "person": person_name,
                    "status": status,
                    "copy_type": copy_type,
                    "engraving_text": engraving_text,
                    "engraving_notes": engraving_notes,
                    "engraving_signature": engraving_signature,
                }
            )
            already_in_catalog[-1]["row"] = row_num
            if sku and matched_item.name.strip().lower() != name.lower():
                sku_name_mismatches.append(
                    {
                        "row": row_num,
                        "import_name": name,
                        "existing_name": matched_item.name,
                        "sku": sku,
                    }
                )
            already_in_catalog[-1].update(
                {
                    "is_sku_unicorn": is_sku_unicorn,
                    "is_variant_unicorn": is_variant_unicorn,
                    "is_edge_unicorn": is_edge_unicorn,
                }
            )
        else:
            if matches_set_sku:
                if person_name and matched_set:
                    for membership in matched_set.members:
                        member_item = membership.item
                        if not member_item:
                            continue
                        accepts_handle_color = accepts_set_handle_variants(
                            member_item.name, member_item.category
                        )
                        member_color = (
                            _resolve_import_variant_color(
                                member_item.name,
                                member_item.category or "",
                                color,
                            )
                            if accepts_handle_color
                            else UNKNOWN_COLOR
                        )
                        member_variant = _find_import_variant(member_item, member_color)
                        member_multiplier = membership.quantity or 1
                        member_quantity_purchased = (
                            quantity_purchased * member_multiplier
                            if quantity_purchased is not None
                            else (member_multiplier if member_multiplier > 1 else None)
                        )
                        member_quantity_given_away = (
                            quantity_given_away * member_multiplier
                            if quantity_given_away is not None
                            else None
                        )
                        ownership_entries.append(
                            {
                                "row": row_num,
                                "person": person_name,
                                "item_name": member_item.name,
                                "sku": member_item.sku,
                                "item_id": member_item.id,
                                "color": member_color,
                                "display_color": _preview_import_color(
                                    member_color,
                                    not accepts_handle_color,
                                ),
                                "status": status,
                                "notes": notes,
                                "non_catalog": not member_item.in_catalog,
                                "availability": member_item.availability,
                                "is_sku_unicorn": False,
                                "is_variant_unicorn": is_variant_unicorn,
                                "is_edge_unicorn": False,
                                "quantity_purchased": member_quantity_purchased,
                                "quantity_given_away": member_quantity_given_away,
                                "copy_type": copy_type,
                                "engraving_text": engraving_text,
                                "engraving_notes": engraving_notes,
                                "engraving_signature": engraving_signature,
                                "is_new_variant": member_variant is None,
                                "is_new_person": person_name.lower()
                                not in existing_persons,
                            }
                        )
                    continue
                bucket = set_sku_collisions
            else:
                bucket = (
                    likely_unicorns
                    if is_sku_unicorn
                    or is_variant_unicorn
                    or is_edge_unicorn
                    or not sku
                    else new_items_list
                )
            bucket.append(
                {
                    "name": name,
                    "sku": sku,
                    "color": target_color,
                    "display_color": _preview_import_color(target_color, is_cookware),
                    "edge_type": edge_type,
                    "non_catalog": non_catalog,
                    "availability": availability,
                    "availability_label": _availability_preview_fields(availability)[0],
                    "availability_badge_class": _availability_preview_fields(
                        availability
                    )[1],
                    "is_sku_unicorn": is_sku_unicorn,
                    "is_variant_unicorn": is_variant_unicorn,
                    "is_edge_unicorn": is_edge_unicorn,
                    "quantity_purchased": quantity_purchased,
                    "quantity_given_away": quantity_given_away,
                    "copy_type": copy_type,
                    "engraving_text": engraving_text,
                    "engraving_notes": engraving_notes,
                    "engraving_signature": engraving_signature,
                    "category": category,
                    "notes": notes,
                    "person": person_name,
                    "status": status,
                    "row": row_num,
                    "matches_set_sku": matches_set_sku,
                    "matched_set_name": (
                        matched_set.name if matched_set is not None else None
                    ),
                }
            )

        if person_name and matched_item:
            person_obj = existing_persons.get(person_name.lower())
            if person_obj:
                if existing_variant:
                    existing_o = Ownership.query.filter_by(
                        person_id=person_obj.id,
                        variant_id=existing_variant.id,
                        copy_type=copy_type,
                        engraving_signature=engraving_signature,
                    ).first()
                    if existing_o:
                        if existing_o.status != status:
                            conflicts.append(
                                {
                                    "row": row_num,
                                    "person": person_name,
                                    "item": matched_item.name,
                                    "sku": matched_item.sku,
                                    "color": color,
                                    "existing_status": existing_o.status,
                                    "import_status": status,
                                    "oid": existing_o.id,
                                }
                            )
                        continue
            ownership_entries.append(
                {
                    "row": row_num,
                    "person": person_name,
                    "item_name": matched_item.name,
                    "sku": matched_item.sku,
                    "item_id": matched_item.id,
                    "color": target_color,
                    "display_color": _preview_import_color(target_color, is_cookware),
                    "status": status,
                    "notes": notes,
                    "non_catalog": non_catalog,
                    "availability": matched_item.availability,
                    "is_sku_unicorn": is_sku_unicorn,
                    "is_variant_unicorn": is_variant_unicorn,
                    "is_edge_unicorn": is_edge_unicorn,
                    "quantity_purchased": quantity_purchased,
                    "quantity_given_away": quantity_given_away,
                    "copy_type": copy_type,
                    "engraving_text": engraving_text,
                    "engraving_notes": engraving_notes,
                    "engraving_signature": engraving_signature,
                    "is_new_variant": existing_variant is None,
                    "is_new_person": person_name.lower() not in existing_persons,
                }
            )

    return render_template(
        "import_preview.html",
        already_in_catalog=already_in_catalog,
        sku_name_mismatches=sku_name_mismatches,
        new_items=new_items_list,
        new_item_groups=_group_import_rows(new_items_list, base_index=0),
        likely_unicorns=likely_unicorns,
        likely_unicorn_groups=_group_import_rows(
            likely_unicorns, base_index=len(new_items_list)
        ),
        set_sku_collisions=set_sku_collisions,
        set_sku_collision_groups=_group_import_rows(
            set_sku_collisions,
            base_index=len(new_items_list) + len(likely_unicorns),
        ),
        set_rows=set_rows,
        ownership_entries=ownership_entries,
        ownership_groups=_group_import_rows(ownership_entries, base_index=0),
        conflicts=conflicts,
        errors=errors,
        total_rows=len(parsed_rows),
        item_rows_total=len(new_items_list)
        + len(likely_unicorns)
        + len(set_sku_collisions),
        set_rows_total=len(set_rows),
        edge_types=EDGE_TYPES,
        edgeless_categories=EDGELESS_CATEGORIES,
        status_options=STATUS_OPTIONS,
        person_override=person_override,
    )


@data_bp.route("/completion-import", methods=["GET", "POST"])
@admin_required
def completion_import_page():
    """Render the completion import page."""
    recent_completion_imports = (
        db.session.execute(
            db.select(ActivityEvent)
            .filter_by(kind="import")
            .where(ActivityEvent.title == "Completion import complete")
            .order_by(desc(ActivityEvent.occurred_at), desc(ActivityEvent.id))
            .limit(5)
        )
        .scalars()
        .all()
    )
    if request.method == "GET":
        return render_template(
            "completion_import.html",
            people=Person.query.order_by(Person.name).all(),
            recent_completion_imports=[
                {
                    "title": event.title,
                    "details": event.details,
                    "time": format_container_time(event.occurred_at),
                }
                for event in recent_completion_imports
            ],
            preview=None,
            export_name=f"cutco_completion_result_{date.today().isoformat()}.csv",
        )

    pasted_rows = request.form.get("rows_text", "")
    uploaded_file = request.files.get("csvfile")

    parsed_rows, parse_error = _read_completion_rows(uploaded_file, pasted_rows)
    if parse_error:
        flash(parse_error, "error")
        return render_template(
            "completion_import.html",
            people=Person.query.order_by(Person.name).all(),
            recent_completion_imports=[
                {
                    "title": event.title,
                    "details": event.details,
                    "time": format_container_time(event.occurred_at),
                }
                for event in recent_completion_imports
            ],
            preview=None,
            export_name=f"cutco_completion_result_{date.today().isoformat()}.csv",
        )

    person_override = request.form.get("person_override", "").strip() or None
    preview = _build_completion_preview(parsed_rows, person_override=person_override)
    return render_template(
        "completion_import_preview.html",
        preview=preview,
        person_override=person_override,
        people=Person.query.order_by(Person.name).all(),
        recent_completion_imports=[
            {
                "title": event.title,
                "details": event.details,
                "time": format_container_time(event.occurred_at),
            }
            for event in recent_completion_imports
        ],
        export_name=f"cutco_completion_result_{date.today().isoformat()}.csv",
    )


@data_bp.route("/completion-import/export", methods=["POST"])
@admin_required
def completion_import_export():
    """Export completion import rows as CSV."""
    export_count = int(request.form.get("export_count", 0) or 0)
    rows = []
    for idx in range(export_count):
        rows.append(
            {
                "person": request.form.get(f"export_person_{idx}", "").strip(),
                "sku": request.form.get(f"export_sku_{idx}", "").strip(),
                "item": request.form.get(f"export_item_{idx}", "").strip(),
                "color": request.form.get(f"export_display_color_{idx}", "").strip()
                or "—",
                "quantity": request.form.get(f"export_quantity_{idx}", "").strip(),
                "action": request.form.get(f"export_action_{idx}", "").strip(),
                "notes": request.form.get(f"export_note_{idx}", "").strip(),
                "source_rows": request.form.get(
                    f"export_source_rows_{idx}", ""
                ).strip(),
            }
        )

    csv_buffer = io.StringIO()
    writer = csv.writer(csv_buffer)
    writer.writerow(
        [
            "person",
            "sku",
            "item",
            "color",
            "total_quantity",
            "action",
            "notes",
            "source_rows",
        ]
    )
    for row in rows:
        writer.writerow(
            [
                row["person"],
                row["sku"],
                row["item"],
                row["color"],
                row["quantity"],
                row["action"],
                row["notes"],
                row["source_rows"],
            ]
        )
    csv_buffer.seek(0)
    filename = _safe_csv_filename(
        request.form.get(
            "filename", f"cutco_completion_result_{date.today().isoformat()}.csv"
        )
    )
    logger.info("Completion export requested: %d rows (%s)", len(rows), filename)
    return Response(
        csv_buffer.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@data_bp.route("/completion-import/missing-export", methods=["POST"])
@admin_required
def completion_import_missing_export():
    """Export unresolved completion rows as CSV."""
    export_count = int(request.form.get("export_count", 0) or 0)
    rows = []
    for idx in range(export_count):
        rows.append(
            {
                "person": request.form.get(f"missing_person_{idx}", "").strip(),
                "missing_sku": request.form.get(f"missing_sku_{idx}", "").strip(),
                "item": request.form.get(f"missing_item_{idx}", "").strip(),
                "category": request.form.get(f"missing_category_{idx}", "").strip()
                or "—",
                "availability": request.form.get(
                    f"missing_availability_{idx}", ""
                ).strip()
                or "public",
            }
        )

    csv_buffer = io.StringIO()
    writer = csv.writer(csv_buffer)
    writer.writerow(["person", "missing_sku", "item", "category", "availability"])
    for row in rows:
        writer.writerow(
            [
                row["person"],
                row["missing_sku"],
                row["item"],
                row["category"],
                row["availability"],
            ]
        )
    csv_buffer.seek(0)
    filename = _safe_csv_filename(
        request.form.get(
            "filename", f"cutco_completion_missing_{date.today().isoformat()}.csv"
        )
    )
    logger.info(
        "Completion missing export requested: %d rows (%s)", len(rows), filename
    )
    return Response(
        csv_buffer.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@data_bp.route("/completion-import/confirm", methods=["POST"])
@admin_required
def completion_import_confirm():
    """Apply a completion import preview."""
    from sqlalchemy.exc import SQLAlchemyError

    existing_persons = {person.name.lower(): person for person in Person.query.all()}
    item_count = int(request.form.get("rolled_count", 0) or 0)
    total_rows = int(request.form.get("total_rows", 0) or 0)
    selected_rows = 0
    processed_rows = 0
    created_ownership = 0
    updated_ownership = 0
    created_people = 0
    skipped_details = []
    export_rows = []

    try:
        for row_index in range(item_count):
            row_num = request.form.get(f"row_input_{row_index}", type=int)
            person_name = request.form.get(f"row_person_{row_index}", "").strip()
            sku = request.form.get(f"row_sku_{row_index}", "").strip()
            item_name = request.form.get(f"row_item_{row_index}", "").strip() or None

            if request.form.get(f"row_accept_{row_index}") != "on":
                _append_import_skip_detail(
                    skipped_details,
                    row_num,
                    "Not selected during import review.",
                    name=item_name,
                    sku=sku,
                )
                continue
            selected_rows += 1

            item_id = int(request.form.get(f"row_item_id_{row_index}", 0) or 0)
            item = db.session.get(Item, item_id)
            if not item:
                _append_import_skip_detail(
                    skipped_details,
                    row_num,
                    "Matched catalog item was not found during confirmation.",
                    name=item_name,
                    sku=sku,
                )
                continue

            quantity, qty_error = parse_positive_whole_number(
                request.form.get(f"row_quantity_{row_index}", "")
            )
            if qty_error:
                _append_import_skip_detail(
                    skipped_details, row_num, qty_error, name=item_name, sku=sku
                )
                continue
            if quantity is None:
                _append_import_skip_detail(
                    skipped_details,
                    row_num,
                    "Quantity is required.",
                    name=item_name,
                    sku=sku,
                )
                continue

            notes = request.form.get(f"row_note_{row_index}", "").strip() or None
            color = (
                request.form.get(f"row_color_{row_index}", "").strip() or UNKNOWN_COLOR
            )
            target_color = color if color and color != UNKNOWN_COLOR else UNKNOWN_COLOR

            person = existing_persons.get(person_name.lower())
            if not person:
                person = Person(name=person_name)
                db.session.add(person)
                db.session.flush()
                existing_persons[person_name.lower()] = person
                created_people += 1

            variant = next(
                (
                    existing_variant
                    for existing_variant in item.variants
                    if existing_variant.color.lower() == target_color.lower()
                ),
                None,
            )
            if not variant:
                variant = ItemVariant(
                    item_id=item.id, color=target_color, source="collection_import"
                )
                db.session.add(variant)
                db.session.flush()

            existing_o = Ownership.query.filter_by(
                person_id=person.id, variant_id=variant.id
            ).first()
            if existing_o:
                existing_o.status = "Owned"
                existing_o.quantity_purchased = (
                    existing_o.quantity_purchased or 0
                ) + quantity
                if notes:
                    existing_o.notes = _merge_note_text(existing_o.notes, notes)
                updated_ownership += 1
                action = "Update ownership"
            else:
                db.session.add(
                    Ownership(
                        person_id=person.id,
                        variant_id=variant.id,
                        status="Owned",
                        quantity_purchased=quantity,
                        notes=notes,
                    )
                )
                created_ownership += 1
                action = "Create ownership"

            db.session.flush()
            reconcile_unknown_variant(item)
            processed_rows += 1
            export_rows.append(
                {
                    "person": person.name,
                    "sku": item.sku or sku,
                    "item": item.name,
                    "display_color": (
                        "—" if target_color == UNKNOWN_COLOR else target_color
                    ),
                    "quantity": quantity,
                    "action": action,
                    "notes": notes or "",
                    "source_rows": str(row_num) if row_num is not None else "",
                }
            )

    except SQLAlchemyError as exc:
        db.session.rollback()
        logger.error("Completion import flush failed: %s", exc)
        flash(
            "Completion import failed — database error during processing. No changes were saved.",
            "error",
        )
        return redirect(url_for("data.completion_import_page"))

    if db_commit(db.session):
        for idx in range(int(request.form.get("unresolved_count", 0) or 0)):
            unresolved_row = request.form.get(f"unresolved_row_{idx}", type=int)
            unresolved_person = (
                request.form.get(f"unresolved_person_{idx}", "").strip() or None
            )
            unresolved_sku = (
                request.form.get(f"unresolved_sku_{idx}", "").strip() or None
            )
            _append_import_skip_detail(
                skipped_details,
                unresolved_row,
                request.form.get(f"unresolved_reason_{idx}", "").strip()
                or "Could not resolve row.",
                name=unresolved_person,
                sku=unresolved_sku,
            )

        skipped_details.sort(
            key=lambda entry: (entry["row"] is None, entry["row"] or 0, entry["label"])
        )
        if created_ownership > updated_ownership:
            outcome_note = "mostly created new ownership entries"
        elif updated_ownership > created_ownership:
            outcome_note = "mostly updated existing ownership entries"
        else:
            outcome_note = "a balanced mix of new and updated ownership entries"
        summary = (
            f"Completion import complete — processed {processed_rows} row{'s' if processed_rows != 1 else ''}, "
            f"created {created_ownership} ownership entr{'ies' if created_ownership != 1 else 'y'}, "
            f"updated {updated_ownership} ownership entr{'ies' if updated_ownership != 1 else 'y'}; "
            f"{outcome_note}."
        )
        missing_export_rows = _build_completion_missing_rows(
            [row["person"] for row in export_rows]
        )
        record_activity(
            "import",
            "Completion import complete",
            f"Processed {processed_rows} rows, created {created_ownership} ownership entries, updated {updated_ownership} ownership entries.",
        )
        db.session.commit()
        return render_template(
            "completion_import_result.html",
            summary=summary,
            total_rows=total_rows,
            selected_rows=selected_rows,
            processed_rows=processed_rows,
            skipped_details=skipped_details,
            created_people=created_people,
            created_ownership=created_ownership,
            updated_ownership=updated_ownership,
            export_rows=export_rows,
            export_name=f"cutco_completion_result_{date.today().isoformat()}.csv",
            missing_export_rows=missing_export_rows,
            missing_export_name=f"cutco_completion_missing_{date.today().isoformat()}.csv",
            missing_people_count=len({row["person"] for row in missing_export_rows}),
            missing_catalog_items_count=len(
                Item.query.filter_by(set_only=False, in_catalog=True).all()
            ),
        )
    return redirect(url_for("data.completion_import_page"))


@data_bp.route("/import/confirm", methods=["POST"])
@admin_required
def import_confirm():
    """Apply an item and ownership import preview."""
    from sqlalchemy.exc import SQLAlchemyError

    added_items = 0
    created_sets = 0
    updated_sets = 0
    added_ownership = 0
    added_persons = 0
    item_rows_selected = 0
    item_rows_imported = 0
    set_rows_selected = 0
    set_rows_imported = 0
    own_rows_selected = 0
    own_rows_imported = 0
    skipped_details = []

    existing_items = _build_item_sku_lookup(Item.query.all())
    existing_set_skus = _build_set_sku_lookup(Set.query.all())
    existing_set_ids = {item_set.id for item_set in Set.query.all()}
    existing_names = _build_item_name_lookup(
        [item for item in Item.query.all() if not normalize_sku_value(item.sku)]
    )
    existing_persons = {person.name.lower(): person for person in Person.query.all()}

    item_count = int(request.form.get("item_count", 0) or 0)
    set_count = int(request.form.get("set_count", 0) or 0)
    own_count = int(request.form.get("own_count", 0) or 0)
    total_rows = int(request.form.get("total_rows", 0) or 0)

    try:
        for row_index in range(item_count):
            row_num = request.form.get(f"item_row_{row_index}", type=int)
            name_hint = request.form.get(f"item_name_{row_index}", "").strip() or None
            sku_hint = normalize_sku_value(
                request.form.get(f"item_sku_{row_index}", "")
            )

            if request.form.get(f"item_accept_{row_index}") != "on":
                _append_import_skip_detail(
                    skipped_details,
                    row_num,
                    "Not selected during import review.",
                    name=name_hint,
                    sku=sku_hint,
                )
                continue
            item_rows_selected += 1

            name = request.form.get(f"item_name_{row_index}", "").strip()
            sku = normalize_sku_value(request.form.get(f"item_sku_{row_index}", ""))
            color = _normalize_import_color(
                request.form.get(f"item_color_{row_index}", "")
            )
            edge_type = request.form.get(f"item_edge_{row_index}", "Unknown")
            availability_raw = request.form.get(
                f"item_availability_{row_index}", ""
            ).strip()
            availability_specified = bool(availability_raw)
            availability = canonicalize_availability(availability_raw)
            non_catalog = request.form.get(f"item_non_catalog_{row_index}") == "on"
            is_sku_unicorn = request.form.get(f"item_sku_unicorn_{row_index}") == "on"
            is_variant_unicorn = (
                request.form.get(f"item_variant_unicorn_{row_index}") == "on"
            )
            is_edge_unicorn = request.form.get(f"item_edge_unicorn_{row_index}") == "on"
            quantity_purchased, qty_error = _read_confirm_quantity_field(
                request.form.get(f"item_quantity_purchased_{row_index}", ""),
                "Quantity Purchased",
            )
            if qty_error:
                _append_import_skip_detail(
                    skipped_details,
                    row_num,
                    qty_error,
                    name=name_hint,
                    sku=sku_hint,
                )
                continue
            quantity_given_away, qty_error = _read_confirm_quantity_field(
                request.form.get(f"item_quantity_given_away_{row_index}", ""),
                "Quantity Given Away",
            )
            if qty_error:
                _append_import_skip_detail(
                    skipped_details,
                    row_num,
                    qty_error,
                    name=name_hint,
                    sku=sku_hint,
                )
                continue
            if availability == "public" and (
                non_catalog or is_sku_unicorn or is_variant_unicorn or is_edge_unicorn
            ):
                availability = "non-catalog"
            non_catalog = (
                non_catalog
                or availability != "public"
                or is_sku_unicorn
                or is_variant_unicorn
                or is_edge_unicorn
            )
            category = canonicalize_category(
                request.form.get(f"item_category_{row_index}", "")
            )
            notes = request.form.get(f"item_notes_{row_index}", "").strip() or None
            person_name = request.form.get(f"item_person_{row_index}", "").strip()
            status = request.form.get(f"item_status_{row_index}", "Owned")
            copy_type, engraving_text, engraving_notes, engraving_signature = (
                _read_engraving_fields(
                    request.form, prefix="item_", suffix=f"_{row_index}"
                )
            )

            if not name:
                _append_import_skip_detail(
                    skipped_details,
                    row_num,
                    "Missing name.",
                    name=name_hint,
                    sku=sku_hint,
                )
                continue

            item = _match_import_item(
                existing_items=existing_items,
                existing_names=existing_names,
                sku=sku,
                name=name,
            )
            if not item:
                item = Item(
                    name=name,
                    sku=sku,
                    category=category,
                    edge_type=edge_type,
                    is_unicorn=is_sku_unicorn,
                    edge_is_unicorn=is_edge_unicorn,
                    availability=availability,
                    in_catalog=availability == "public" and not non_catalog,
                    notes=notes,
                )
                db.session.add(item)
                db.session.flush()
                if sku:
                    existing_items[sku] = item
                if not sku:
                    existing_names.setdefault(
                        _normalize_variant_lookup_name(name), item
                    )
                added_items += 1
            else:
                if availability_specified or non_catalog:
                    item.availability = availability
                    item.in_catalog = availability == "public" and not item.set_only
                if is_sku_unicorn and not item.is_unicorn:
                    item.is_unicorn = True
                if is_edge_unicorn and not item.edge_is_unicorn:
                    item.edge_is_unicorn = True
                if is_edgeless_category(item.category):
                    item.edge_type = "N/A"
                    item.edge_is_unicorn = False
                if non_catalog:
                    item.in_catalog = False

            target_color = _resolve_import_variant_color(
                item.name, item.category or category, color
            )
            variant = _find_import_variant(item, target_color)
            if not variant:
                variant = ItemVariant(
                    item_id=item.id,
                    color=target_color,
                    is_unicorn=is_variant_unicorn,
                    source="catalog_sync",
                )
                db.session.add(variant)
                db.session.flush()
            elif is_variant_unicorn and not variant.is_unicorn:
                variant.is_unicorn = True

            person = None
            if person_name:
                person = existing_persons.get(person_name.lower())
                if not person:
                    person = Person(name=person_name)
                    db.session.add(person)
                    db.session.flush()
                    existing_persons[person_name.lower()] = person
                    added_persons += 1
                existing_o = Ownership.query.filter_by(
                    person_id=person.id,
                    variant_id=variant.id,
                    copy_type=copy_type,
                    engraving_signature=engraving_signature,
                ).first()
                if existing_o:
                    if existing_o.status != status:
                        continue
                    _add_import_ownership_quantities(
                        existing_o,
                        status=status,
                        notes=existing_o.notes,
                        quantity_purchased=quantity_purchased,
                        quantity_given_away=quantity_given_away,
                        copy_type=copy_type,
                        engraving_text=engraving_text,
                        engraving_notes=engraving_notes,
                        engraving_signature=engraving_signature,
                    )
                else:
                    db.session.add(
                        Ownership(
                            person_id=person.id,
                            variant_id=variant.id,
                            status=status,
                            quantity_purchased=quantity_purchased,
                            quantity_given_away=quantity_given_away,
                            copy_type=copy_type,
                            engraving_text=engraving_text,
                            engraving_notes=engraving_notes,
                            engraving_signature=engraving_signature,
                        )
                    )
                    added_ownership += 1

            db.session.flush()
            reconcile_unknown_variant(item)
            item_rows_imported += 1

        for row_index in range(set_count):
            row_num = request.form.get(f"set_row_{row_index}", type=int)
            set_name_hint = (
                request.form.get(f"set_name_{row_index}", "").strip() or None
            )
            set_sku_hint = normalize_sku_value(
                request.form.get(f"set_sku_{row_index}", "")
            )

            if request.form.get(f"set_accept_{row_index}") != "on":
                _append_import_skip_detail(
                    skipped_details,
                    row_num,
                    "Not selected during import review.",
                    name=set_name_hint,
                    sku=set_sku_hint,
                )
                continue
            set_rows_selected += 1

            set_name = request.form.get(f"set_name_{row_index}", "").strip()
            set_sku = normalize_sku_value(request.form.get(f"set_sku_{row_index}", ""))
            set_member_skus = _parse_set_members_field(
                request.form.get(f"set_members_{row_index}", "")
            )
            if not set_name:
                _append_import_skip_detail(
                    skipped_details,
                    row_num,
                    "Missing set name.",
                    name=set_name_hint,
                    sku=set_sku_hint,
                )
                continue
            if not set_sku:
                _append_import_skip_detail(
                    skipped_details,
                    row_num,
                    "Missing set SKU.",
                    name=set_name_hint,
                    sku=set_sku_hint,
                )
                continue
            if not set_member_skus:
                _append_import_skip_detail(
                    skipped_details,
                    row_num,
                    "Missing set member SKUs.",
                    name=set_name_hint,
                    sku=set_sku_hint,
                )
                continue

            member_counts: dict[str, int] = {}
            for member_sku in set_member_skus:
                member_counts[member_sku] = member_counts.get(member_sku, 0) + 1
            member_entries: list[SetMemberEntry] = []
            missing_members: list[str] = []
            for member_sku, qty in member_counts.items():
                item = existing_items.get(member_sku)
                if not item:
                    missing_members.append(member_sku)
                    continue
                member_entries.append(
                    {"sku": member_sku, "quantity": qty, "name": item.name}
                )
            if missing_members:
                _append_import_skip_detail(
                    skipped_details,
                    row_num,
                    "Missing set member SKU(s): " + ", ".join(missing_members),
                    name=set_name_hint,
                    sku=set_sku_hint,
                )
                continue

            existing_set_by_name = Set.query.filter(
                db.func.lower(Set.name) == set_name.lower()
            ).first()
            existing_set_by_sku = existing_set_skus.get(set_sku)
            if (
                existing_set_by_name
                and existing_set_by_sku
                and existing_set_by_name.id != existing_set_by_sku.id
            ):
                _append_import_skip_detail(
                    skipped_details,
                    row_num,
                    "Set name and set SKU point to different existing sets.",
                    name=set_name,
                    sku=set_sku,
                )
                continue

            item_set = existing_set_by_sku or existing_set_by_name
            if item_set is None:
                item_set = get_or_create_set(set_name)
            if item_set.id in existing_set_ids:
                updated_sets += 1
            else:
                created_sets += 1
                existing_set_ids.add(item_set.id)
            if item_set.name != set_name:
                item_set.name = set_name
            if item_set.sku != set_sku:
                item_set.sku = set_sku
                if set_sku is not None:
                    existing_set_skus[set_sku] = item_set
            item_set.member_data = json.dumps(member_entries, ensure_ascii=False)
            existing_members = {member.item_id: member for member in item_set.members}
            incoming_member_ids: set[int] = set()
            for member in member_entries:
                member_sku = member["sku"]
                item = existing_items.get(member_sku)
                if not item:
                    continue
                qty = member["quantity"]
                if item.id not in existing_members:
                    db.session.add(
                        ItemSetMember(set_id=item_set.id, item_id=item.id, quantity=qty)
                    )
                else:
                    existing_members[item.id].quantity = qty
                incoming_member_ids.add(item.id)
            for membership in list(item_set.members):
                if membership.item_id not in incoming_member_ids:
                    db.session.delete(membership)
            set_rows_imported += 1

        for row_index in range(own_count):
            row_num = request.form.get(f"own_row_{row_index}", type=int)
            item_name_hint = (
                request.form.get(f"own_item_name_{row_index}", "").strip() or None
            )
            sku_hint = normalize_sku_value(
                request.form.get(f"own_item_sku_{row_index}", "")
            )

            if request.form.get(f"own_accept_{row_index}") != "on":
                _append_import_skip_detail(
                    skipped_details,
                    row_num,
                    "Not selected during import review.",
                    name=item_name_hint,
                    sku=sku_hint,
                )
                continue
            own_rows_selected += 1

            item_id = int(request.form.get(f"own_item_id_{row_index}", 0))
            person_name = request.form.get(f"own_person_{row_index}", "").strip()
            color = _normalize_import_color(
                request.form.get(f"own_color_{row_index}", "")
            )
            status = request.form.get(f"own_status_{row_index}", "Owned")
            notes = request.form.get(f"own_notes_{row_index}", "").strip() or None
            copy_type, engraving_text, engraving_notes, engraving_signature = (
                _read_engraving_fields(
                    request.form, prefix="own_", suffix=f"_{row_index}"
                )
            )
            quantity_purchased, qty_error = _read_confirm_quantity_field(
                request.form.get(f"own_quantity_purchased_{row_index}", ""),
                "Quantity Purchased",
            )
            if qty_error:
                _append_import_skip_detail(
                    skipped_details,
                    row_num,
                    qty_error,
                    name=item_name_hint,
                    sku=sku_hint,
                )
                continue
            quantity_given_away, qty_error = _read_confirm_quantity_field(
                request.form.get(f"own_quantity_given_away_{row_index}", ""),
                "Quantity Given Away",
            )
            if qty_error:
                _append_import_skip_detail(
                    skipped_details,
                    row_num,
                    qty_error,
                    name=item_name_hint,
                    sku=sku_hint,
                )
                continue
            is_sku_unicorn = request.form.get(f"own_sku_unicorn_{row_index}") == "on"
            is_variant_unicorn = (
                request.form.get(f"own_variant_unicorn_{row_index}") == "on"
            )
            is_edge_unicorn = request.form.get(f"own_edge_unicorn_{row_index}") == "on"

            item = db.session.get(Item, item_id)
            if not item:
                _append_import_skip_detail(
                    skipped_details,
                    row_num,
                    "Matched catalog item was not found during confirmation.",
                    name=item_name_hint,
                    sku=sku_hint,
                )
                continue
            if not person_name:
                _append_import_skip_detail(
                    skipped_details,
                    row_num,
                    "Missing person/collector name.",
                    name=item_name_hint,
                    sku=sku_hint,
                )
                continue
            if is_sku_unicorn and not item.is_unicorn:
                item.is_unicorn = True
            if is_edge_unicorn and not item.edge_is_unicorn:
                item.edge_is_unicorn = True
            if is_edgeless_category(item.category):
                item.edge_type = "N/A"
                item.edge_is_unicorn = False

            person = existing_persons.get(person_name.lower())
            if not person:
                person = Person(name=person_name)
                db.session.add(person)
                db.session.flush()
                existing_persons[person_name.lower()] = person
                added_persons += 1

            target_color = _resolve_import_variant_color(
                item.name, item.category or "", color
            )
            variant = _find_import_variant(item, target_color)
            if not variant:
                variant = ItemVariant(
                    item_id=item.id,
                    color=target_color,
                    is_unicorn=is_variant_unicorn,
                    source="collection_import",
                )
                db.session.add(variant)
                db.session.flush()
            elif is_variant_unicorn and not variant.is_unicorn:
                variant.is_unicorn = True

            existing_o = Ownership.query.filter_by(
                person_id=person.id,
                variant_id=variant.id,
                copy_type=copy_type,
                engraving_signature=engraving_signature,
            ).first()
            if existing_o:
                if existing_o.status != status:
                    continue
                _merge_import_ownership(
                    existing_o,
                    status=status,
                    notes=notes,
                    quantity_purchased=quantity_purchased,
                    quantity_given_away=quantity_given_away,
                    copy_type=copy_type,
                    engraving_text=engraving_text,
                    engraving_notes=engraving_notes,
                    engraving_signature=engraving_signature,
                )
            else:
                db.session.add(
                    Ownership(
                        person_id=person.id,
                        variant_id=variant.id,
                        status=status,
                        notes=notes,
                        quantity_purchased=quantity_purchased,
                        quantity_given_away=quantity_given_away,
                        copy_type=copy_type,
                        engraving_text=engraving_text,
                        engraving_notes=engraving_notes,
                        engraving_signature=engraving_signature,
                    )
                )
                added_ownership += 1

            db.session.flush()
            reconcile_unknown_variant(item)
            own_rows_imported += 1

    except SQLAlchemyError as exc:
        db.session.rollback()
        logger.error("Import flush failed: %s", exc)
        flash(
            "Import failed — database error during processing. No changes were saved.",
            "error",
        )
        return redirect(url_for("catalog.catalog"))

    if db_commit(db.session):
        logger.info(
            "Import complete: %d items, %d ownership, %d persons",
            added_items,
            added_ownership,
            added_persons,
        )
        selected_rows = item_rows_selected + set_rows_selected + own_rows_selected
        imported_rows = item_rows_imported + set_rows_imported + own_rows_imported
        error_count = int(request.form.get("error_count", 0) or 0)
        for idx in range(error_count):
            row_num = request.form.get(f"error_row_{idx}", type=int)
            name_hint = request.form.get(f"error_name_{idx}", "").strip() or None
            sku_hint = request.form.get(f"error_sku_{idx}", "").strip().upper() or None
            reason = (
                request.form.get(f"error_reason_{idx}", "").strip()
                or "Could not parse row."
            )
            _append_import_skip_detail(
                skipped_details,
                row_num,
                reason,
                name=name_hint,
                sku=sku_hint,
            )

        conflict_count = int(request.form.get("conflict_count", 0) or 0)
        for idx in range(conflict_count):
            row_num = request.form.get(f"conflict_row_{idx}", type=int)
            item_name = request.form.get(f"conflict_item_{idx}", "").strip() or None
            sku_hint = (
                request.form.get(f"conflict_sku_{idx}", "").strip().upper() or None
            )
            person_name = request.form.get(f"conflict_person_{idx}", "").strip()
            existing_status = request.form.get(
                f"conflict_existing_status_{idx}", ""
            ).strip()
            import_status = request.form.get(
                f"conflict_import_status_{idx}", ""
            ).strip()
            reason = (
                f"Existing entry for {person_name or 'collector'} kept unchanged "
                f"({existing_status or 'existing'} vs {import_status or 'import'})."
            )
            _append_import_skip_detail(
                skipped_details,
                row_num,
                reason,
                name=item_name,
                sku=sku_hint,
            )

        skipped_details.sort(
            key=lambda entry: (entry["row"] is None, entry["row"] or 0, entry["label"])
        )
        parts = []
        if total_rows:
            parts.append(f"read {total_rows} row{'s' if total_rows != 1 else ''}")
        parts.append(f"selected {selected_rows} row{'s' if selected_rows != 1 else ''}")
        parts.append(f"imported {imported_rows} row{'s' if imported_rows != 1 else ''}")
        if added_items:
            parts.append(f"{added_items} item{'s' if added_items != 1 else ''}")
        if added_persons:
            parts.append(
                f"{added_persons} collector{'s' if added_persons != 1 else ''}"
            )
        if created_sets:
            parts.append(
                f"{created_sets} set{'s' if created_sets != 1 else ''} created"
            )
        if updated_sets:
            parts.append(
                f"{updated_sets} set{'s' if updated_sets != 1 else ''} updated"
            )
        if added_ownership:
            parts.append(
                f"{added_ownership} ownership entr{'ies' if added_ownership != 1 else 'y'}"
            )
        summary = (
            "Import complete — added "
            + (", ".join(parts) if parts else "nothing new")
            + "."
        )
        record_activity(
            "import",
            "Import complete",
            f"Imported {imported_rows} rows, added {added_items} items, created {created_sets} sets, updated {updated_sets} sets, {added_persons} collectors, {added_ownership} ownership entries.",
        )
        db.session.commit()
        return render_template(
            "import_result.html",
            summary=summary,
            total_rows=total_rows,
            selected_rows=selected_rows,
            imported_rows=imported_rows,
            skipped_details=skipped_details,
            added_items=added_items,
            created_sets=created_sets,
            updated_sets=updated_sets,
            added_persons=added_persons,
            added_ownership=added_ownership,
        )
    return redirect(url_for("catalog.catalog"))


import blueprints.data_variant_sync  # noqa: E402,F401
