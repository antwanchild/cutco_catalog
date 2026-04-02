import csv
import io
import json
import logging
import os
import re
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date
from logging.handlers import RotatingFileHandler

import openpyxl
import requests
from bs4 import BeautifulSoup
from flask import (Flask, flash, jsonify, redirect, render_template,
                   request, Response, url_for)
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy.orm import selectinload

# ── Logging ───────────────────────────────────────────────────────────────────

LOG_LEVEL   = os.environ.get("LOG_LEVEL", "INFO").upper()
LOG_DIR     = os.environ.get("LOG_DIR", "/data/logs")
_log_level  = getattr(logging, LOG_LEVEL, logging.INFO)
_log_format = logging.Formatter(
    "[%(asctime)s] [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S %z",
)

# Console handler — always on
_console_handler = logging.StreamHandler()
_console_handler.setFormatter(_log_format)

logging.basicConfig(level=_log_level, handlers=[_console_handler])
logger = logging.getLogger(__name__)

# Rotating file handler — writes to /data/logs/cutco.log
# 5 MB per file, keep 5 rotated files (25 MB max)
try:
    os.makedirs(LOG_DIR, exist_ok=True)
    _file_handler = RotatingFileHandler(
        os.path.join(LOG_DIR, "cutco.log"),
        maxBytes=5 * 1024 * 1024,
        backupCount=5,
    )
    _file_handler.setFormatter(_log_format)
    logging.getLogger().addHandler(_file_handler)
    logger.info("File logging active: %s", LOG_DIR)
except OSError as exc:
    logger.warning("Could not set up file logging (%s) — console only", exc)

# ── App / DB ──────────────────────────────────────────────────────────────────

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "cutco-vault-dev-key")
app.config["SQLALCHEMY_DATABASE_URI"] = os.environ.get(
    "DATABASE_URL", "sqlite:////data/cutco.db")
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

db = SQLAlchemy(app)

# ── Constants ─────────────────────────────────────────────────────────────────

EDGE_TYPES = ["Straight", "Double-D", "Serrated", "Micro-D", "Tec Edge", "Unknown"]
STATUS_OPTIONS = ["Owned", "Wishlist", "Sold", "Traded"]
ADMIN_TOKEN             = os.environ.get("ADMIN_TOKEN", "admin")
DATA_DIR                = os.environ.get("DATA_DIR", "/data")
DISCORD_WEBHOOK_URL     = os.environ.get("DISCORD_WEBHOOK_URL", "").strip()
SHARPEN_METHODS         = ["Home Sharpener", "Whetstone", "Cutco Service", "Professional", "Other"]
SHARPEN_THRESHOLD_DAYS  = int(os.environ.get("SHARPEN_THRESHOLD_DAYS", "180"))
BAKEWARE_THRESHOLD_DAYS = int(os.environ.get("BAKEWARE_THRESHOLD_DAYS", "60"))
_bakeware_env = os.environ.get("BAKEWARE_CATEGORIES", "Ake Cookware,Cookware,Bakeware")
BAKEWARE_CATEGORIES     = {cat.strip() for cat in _bakeware_env.split(",") if cat.strip()}
UNKNOWN_COLOR = "Unknown / Unspecified"
APP_VERSION = os.environ.get("APP_VERSION", "dev")

SCRAPE_CATEGORIES = [
    ("Utility Knives",  "https://www.cutco.com/shop/utility-knives"),
    ("Chef Knives",     "https://www.cutco.com/shop/chef-knives"),
    ("Paring Knives",   "https://www.cutco.com/shop/paring-knives"),
    ("Outdoor Knives",  "https://www.cutco.com/shop/outdoor-knives"),
    ("Everyday Knives", "https://www.cutco.com/shop/everyday-knives"),
    ("Table Knives",    "https://www.cutco.com/shop/table-knives"),
    ("Steak Knives",    "https://www.cutco.com/shop/steak-knives"),
    ("Kitchen Tools",   "https://www.cutco.com/shop/kitchen-tools"),
    ("Gadgets",         "https://www.cutco.com/shop/gadgets"),
    ("Cutting Boards",  "https://www.cutco.com/shop/cutting-boards"),
    ("Accessories",     "https://www.cutco.com/shop/cooks-tools"),
    ("Flatware",         "https://www.cutco.com/shop/flatware"),
    ("Tableware",        "https://www.cutco.com/shop/tableware"),
    ("Cookware",         "https://www.cutco.com/shop/cookware"),
    ("Ake Cookware",     "https://www.cutco.com/shop/ake-cookware"),
    ("Storage",          "https://www.cutco.com/shop/storage"),
    ("Sheaths",    "https://www.cutco.com/shop/kitchen-knife-sheaths"),
    ("Garden Tools",     "https://www.cutco.com/shop/garden-tools"),
    ("Kitchen Knives",   "https://www.cutco.com/shop/kitchen-knives"),
]

_BUNDLE_KEYWORDS = {"gift", "additional"}

# Override the scraped category for specific SKUs.
# Useful when a product appears on a category page that doesn't reflect its
# true type (e.g. the Shears Holster is on Accessories but belongs with sheaths).
CATEGORY_OVERRIDES: dict[str, str] = {
    "79": "Sheaths",  # Shears Holster
}


def _resolve_category(sku: str, scraped_category: str, name: str = "") -> str:
    """Return the effective category for an item, applying overrides."""
    if sku in CATEGORY_OVERRIDES:
        return CATEGORY_OVERRIDES[sku]
    # SKUs ending in -N (e.g. "4135-2") are sheath/accessory variants
    if re.search(r"-\d+$", sku):
        return "Sheaths"
    # Items whose name contains "sheath" (but not "with sheath", e.g. "Cleaver with Sheath")
    # belong with sheaths regardless of which category page they were discovered on
    if "sheath" in name.lower() and "with sheath" not in name.lower():
        return "Sheaths"
    return scraped_category

# Words that indicate a product is a bundle/set, not a standalone catalog item.
# Knife blocks (e.g. "Gourmet Set Block") are excluded from this check.
_SET_NAME_PATTERN = re.compile(
    r"\b(set|setting|pack|mates|classics|combo|collection|favorites|starters|bundle|companions|gift\s+box)\b",
    re.IGNORECASE,
)


def _is_set_product(name: str) -> bool:
    """Return True if the name suggests a bundle/set rather than a single item."""
    if not name or not _SET_NAME_PATTERN.search(name):
        return False
    # Knife storage blocks are standalone items — their name contains "Set Block"
    # as a phrase (e.g. "Gourmet Set Block"), not "Set with Block".
    if "set block" in name.lower():
        return False
    return True


SCRAPE_SETS_URL = "https://www.cutco.com/shop/knife-sets"
SCRAPE_HEADERS  = {"User-Agent": "Mozilla/5.0 (compatible; CutcoVaultBot/1.0)"}
REQUEST_TIMEOUT = 12  # seconds for all outbound HTTP requests

# Default: nothing blocked — all categories shown in preview before import.
# Override via env: SYNC_BLOCKED_CATEGORIES="Tableware,Accessories"
_blocked_env = os.environ.get("SYNC_BLOCKED_CATEGORIES", "")
SYNC_BLOCKED_CATEGORIES = {cat_name.strip() for cat_name in _blocked_env.split(",") if cat_name.strip()}

# Set column names as they appear in the spreadsheet
SPREADSHEET_SET_COLUMNS = [
    "Beast", "Fanatic", "SIGNATURE", "HOMEMAKER",
    "Accomplished Chef", "CUTCO Kitchen", "BEAST2", "HOMEMAKER2",
    "Accomplished Chef3", "CUTCO Kitchen4",
]

# ── Models ────────────────────────────────────────────────────────────────────

# Many-to-many join table: items <-> sets
item_sets = db.Table(
    "item_sets",
    db.Column("item_id", db.Integer, db.ForeignKey("items.id"), primary_key=True),
    db.Column("set_id",  db.Integer, db.ForeignKey("sets.id"),  primary_key=True),
)


class Item(db.Model):
    __tablename__ = "items"

    id         = db.Column(db.Integer, primary_key=True)
    name       = db.Column(db.String(160), nullable=False)
    sku        = db.Column(db.String(60),  nullable=True, unique=True)
    category   = db.Column(db.String(80),  nullable=True)
    edge_type  = db.Column(db.String(40),  nullable=False, default="Unknown")
    is_unicorn = db.Column(db.Boolean,     nullable=False, default=False)
    in_catalog = db.Column(db.Boolean,     nullable=False, default=True)
    cutco_url  = db.Column(db.String(300), nullable=True)
    msrp       = db.Column(db.Float,       nullable=True)
    notes      = db.Column(db.Text,        nullable=True)

    variants = db.relationship("ItemVariant", backref="item",
                               lazy=True, cascade="all, delete-orphan",
                               order_by="ItemVariant.color")
    sets     = db.relationship("Set", secondary=item_sets,
                               back_populates="items", lazy="select")

    @property
    def any_unicorn(self) -> bool:
        """True if the item itself is flagged unicorn OR any specific variant is."""
        return self.is_unicorn or any(variant.is_unicorn for variant in self.variants)


class ItemVariant(db.Model):
    __tablename__ = "item_variants"

    id         = db.Column(db.Integer, primary_key=True)
    item_id    = db.Column(db.Integer, db.ForeignKey("items.id"), nullable=False)
    color      = db.Column(db.String(80), nullable=False, default=UNKNOWN_COLOR)
    is_unicorn = db.Column(db.Boolean,    nullable=False, default=False)
    notes      = db.Column(db.Text, nullable=True)

    ownerships = db.relationship("Ownership", backref="variant",
                                 lazy=True, cascade="all, delete-orphan")


class Set(db.Model):
    __tablename__ = "sets"

    id    = db.Column(db.Integer, primary_key=True)
    name  = db.Column(db.String(120), nullable=False, unique=True)
    sku   = db.Column(db.String(20),  nullable=True)
    notes = db.Column(db.Text, nullable=True)

    items = db.relationship("Item", secondary=item_sets,
                            back_populates="sets", lazy="select")


class Person(db.Model):
    __tablename__ = "people"

    id    = db.Column(db.Integer, primary_key=True)
    name  = db.Column(db.String(120), nullable=False)
    notes = db.Column(db.Text, nullable=True)

    ownerships = db.relationship("Ownership", backref="person",
                                 lazy=True, cascade="all, delete-orphan")


class Ownership(db.Model):
    __tablename__ = "ownership"

    id           = db.Column(db.Integer, primary_key=True)
    variant_id   = db.Column(db.Integer, db.ForeignKey("item_variants.id"), nullable=False)
    person_id    = db.Column(db.Integer, db.ForeignKey("people.id"),        nullable=False)
    status       = db.Column(db.String(20), nullable=False, default="Owned")
    target_price = db.Column(db.Float, nullable=True)
    notes        = db.Column(db.Text, nullable=True)

    __table_args__ = (db.UniqueConstraint("variant_id", "person_id",
                                          name="uq_variant_person"),)


class BakewareSession(db.Model):
    __tablename__ = "bakeware_sessions"

    id         = db.Column(db.Integer, primary_key=True)
    item_id    = db.Column(db.Integer, db.ForeignKey("items.id"), nullable=False)
    baked_on   = db.Column(db.String(10), nullable=False)   # ISO date YYYY-MM-DD
    what_made  = db.Column(db.String(200), nullable=False)
    rating     = db.Column(db.Integer, nullable=True)        # 1–5
    notes      = db.Column(db.Text, nullable=True)

    item = db.relationship("Item", backref=db.backref(
        "bakeware_sessions", lazy=True,
        order_by="BakewareSession.baked_on.desc()",
    ))


class SharpeningLog(db.Model):
    __tablename__ = "sharpening_log"

    id           = db.Column(db.Integer, primary_key=True)
    item_id      = db.Column(db.Integer, db.ForeignKey("items.id"), nullable=False)
    sharpened_on = db.Column(db.String(10), nullable=False)   # ISO date YYYY-MM-DD
    method       = db.Column(db.String(60), nullable=False, default="Home Sharpener")
    notes        = db.Column(db.Text, nullable=True)

    item = db.relationship("Item", backref=db.backref(
        "sharpening_log", lazy=True,
        order_by="SharpeningLog.sharpened_on.desc()",
    ))


# ── DB init ───────────────────────────────────────────────────────────────────

def ensure_unknown_variant(item):
    """Guarantee every item has an 'Unknown / Unspecified' color variant.

    This catch-all variant lets ownership be recorded even when a specific
    handle color is not known.  Called after every new Item insert.
    """
    if not any(v.color == UNKNOWN_COLOR for v in item.variants):
        db.session.add(ItemVariant(item_id=item.id, color=UNKNOWN_COLOR))
        db.session.flush()


with app.app_context():
    db.create_all()
    # Incremental migrations for columns added after initial schema
    with db.engine.connect() as _conn:
        for _stmt in [
            "ALTER TABLE sets ADD COLUMN sku VARCHAR(20)",
            "ALTER TABLE item_variants ADD COLUMN is_unicorn BOOLEAN NOT NULL DEFAULT 0",
            "ALTER TABLE items ADD COLUMN msrp REAL",
            "ALTER TABLE ownership ADD COLUMN target_price REAL",
        ]:
            try:
                _conn.execute(db.text(_stmt))
                _conn.commit()
            except Exception:
                pass  # column already exists
    # Remove items with single-digit SKUs — these were erroneously extracted
    # from slug-prefixed URLs like /p/3-inch-gourmet-paring-knife (SKU "3").
    _bad = Item.query.filter(Item.sku.op("GLOB")("[0-9]")).all()
    for _item in _bad:
        logger.info("Removing item with invalid single-digit SKU: %s (sku=%s)", _item.name, _item.sku)
        db.session.delete(_item)
    for item in Item.query.all():
        ensure_unknown_variant(item)
    db.session.commit()
    logger.info("Database ready")

# ── Helpers ───────────────────────────────────────────────────────────────────

def is_admin():
    return request.cookies.get("admin_token") == ADMIN_TOKEN


def _notify_discord(message: str) -> bool:
    """POST a message to the configured Discord webhook. Returns True on success."""
    if not DISCORD_WEBHOOK_URL:
        return False
    try:
        resp = requests.post(DISCORD_WEBHOOK_URL, json={"content": message}, timeout=10)
        resp.raise_for_status()
        logger.info("Discord notification sent (%d chars)", len(message))
        return True
    except Exception as exc:
        logger.warning("Discord notification failed: %s", exc)
        return False


def check_wishlist_targets() -> list[dict]:
    """Return wishlist entries where current MSRP is at or below the target price."""
    hits = []
    entries = (Ownership.query
               .filter_by(status="Wishlist")
               .filter(Ownership.target_price.isnot(None))
               .all())
    for entry in entries:
        msrp = entry.variant.item.msrp
        if msrp is not None and msrp <= entry.target_price:
            hits.append(dict(
                person  = entry.person.name,
                item    = entry.variant.item.name,
                sku     = entry.variant.item.sku,
                target  = entry.target_price,
                msrp    = msrp,
                savings = entry.target_price - msrp,
            ))
    return hits


# ── MSRP diff (web UI) ────────────────────────────────────────────────────────

_MSRP_JOB_FILE  = os.path.join(DATA_DIR, "msrp_job.json")
_msrp_write_lock = threading.Lock()


def _read_msrp_job() -> dict:
    try:
        with open(_MSRP_JOB_FILE) as fh:
            return json.load(fh)
    except (FileNotFoundError, json.JSONDecodeError, OSError):
        return {"status": "idle", "progress": [], "results": None,
                "error": None, "started_at": None, "finished_at": None,
                "update_db": False}


def _write_msrp_job(data: dict) -> None:
    with _msrp_write_lock:
        os.makedirs(DATA_DIR, exist_ok=True)
        tmp = _MSRP_JOB_FILE + ".tmp"
        with open(tmp, "w") as fh:
            json.dump(data, fh)
        os.replace(tmp, _MSRP_JOB_FILE)


def _scrape_price_from_page(url: str) -> float | None:
    """Return the price from a Cutco product page, or None if not found."""
    try:
        resp = requests.get(url, headers=SCRAPE_HEADERS, timeout=REQUEST_TIMEOUT)
        if resp.status_code != 200:
            return None
        soup = BeautifulSoup(resp.text, "html.parser")

        for ld_tag in soup.find_all("script", type="application/ld+json"):
            try:
                data = json.loads(ld_tag.string or "")
                entries = data if isinstance(data, list) else [data]
                for entry in entries:
                    if not isinstance(entry, dict) or entry.get("@type") != "Product":
                        continue
                    offers = entry.get("offers", {})
                    if isinstance(offers, list):
                        offers = offers[0] if offers else {}
                    price_val = offers.get("price") if isinstance(offers, dict) else None
                    if price_val is not None:
                        return float(price_val)
            except (json.JSONDecodeError, ValueError, AttributeError):
                pass

        og_tag = soup.find("meta", property="og:price:amount")
        if og_tag and og_tag.get("content", "").strip():
            try:
                return float(og_tag["content"].replace(",", ""))
            except ValueError:
                pass

        price_el = soup.find(attrs={"itemprop": "price"})
        if price_el:
            raw = price_el.get("content") or price_el.get_text(strip=True)
            price_match = re.search(r"[\d,]+\.?\d*", raw)
            if price_match:
                try:
                    return float(price_match.group().replace(",", ""))
                except ValueError:
                    pass

        for noise in soup.find_all(["script", "style"]):
            noise.decompose()
        dollar_match = re.search(r"\$\s*([\d,]+\.\d{2})", soup.get_text(" ", strip=True))
        if dollar_match:
            try:
                return float(dollar_match.group(1).replace(",", ""))
            except ValueError:
                pass
        return None
    except Exception:
        return None


def _build_msrp_diff(db_items: list, live: dict) -> dict:
    """Diff DB MSRP prices against live scraped prices."""
    db_by_sku = {item.sku: item for item in db_items if item.sku}
    live_skus  = set(live.keys())
    db_skus    = set(db_by_sku.keys())
    result: dict[str, list] = {
        "new": [], "removed": [], "increased": [],
        "decreased": [], "unchanged": [], "no_price": [],
    }
    for sku in sorted(live_skus - db_skus):
        info = live[sku]
        result["new"].append({"sku": sku, "name": info["name"],
                               "db_price": None, "live_price": info["price"]})
    for sku in sorted(db_skus - live_skus):
        item = db_by_sku[sku]
        result["removed"].append({"sku": sku, "name": item.name,
                                   "db_price": item.msrp, "live_price": None})
    for sku in sorted(db_skus & live_skus):
        item       = db_by_sku[sku]
        db_price   = item.msrp
        live_price = live[sku]["price"]
        if live_price is None:
            result["no_price"].append({"sku": sku, "name": item.name,
                                        "db_price": db_price, "live_price": None})
        elif db_price is None:
            result["unchanged"].append({"sku": sku, "name": item.name,
                                         "db_price": None, "live_price": live_price})
        elif live_price > db_price + 0.005:
            result["increased"].append({"sku": sku, "name": item.name,
                                         "db_price": db_price, "live_price": live_price,
                                         "delta": live_price - db_price})
        elif live_price < db_price - 0.005:
            result["decreased"].append({"sku": sku, "name": item.name,
                                         "db_price": db_price, "live_price": live_price,
                                         "delta": live_price - db_price})
        else:
            result["unchanged"].append({"sku": sku, "name": item.name,
                                         "db_price": db_price, "live_price": live_price})
    return result


def _run_msrp_diff_job(update_db: bool) -> None:
    """Background thread: scrape prices, diff against DB, optionally update."""

    def log(msg: str) -> None:
        job = _read_msrp_job()
        job["progress"].append(msg)
        _write_msrp_job(job)

    try:
        with app.app_context():
            log("Scraping live catalog…")
            live_items, _ = scrape_catalog()
            log(f"Found {len(live_items)} items on cutco.com")

            by_sku: dict[str, dict] = {}
            for live_item in live_items:
                sku = live_item.get("sku")
                if sku and sku not in by_sku:
                    by_sku[sku] = {"name": live_item["name"],
                                   "url": live_item["url"], "price": None}

            log(f"Fetching prices for {len(by_sku)} unique SKUs…")
            fetched = 0
            with ThreadPoolExecutor(max_workers=8) as pool:
                future_map = {
                    pool.submit(_scrape_price_from_page, info["url"]): sku
                    for sku, info in by_sku.items() if info.get("url")
                }
                for future in as_completed(future_map):
                    sku = future_map[future]
                    by_sku[sku]["price"] = future.result()
                    fetched += 1
                    if fetched % 20 == 0:
                        log(f"  …{fetched}/{len(future_map)} prices fetched")

            priced = sum(1 for info in by_sku.values() if info["price"] is not None)
            log(f"Prices found: {priced}/{len(by_sku)}")

            db_items = Item.query.filter(Item.sku.isnot(None)).all()
            log(f"Loaded {len(db_items)} DB items — building diff…")
            diff = _build_msrp_diff(db_items, by_sku)

            changes = len(diff["increased"]) + len(diff["decreased"])
            log(f"Done — {changes} price change(s), {len(diff['new'])} new, "
                f"{len(diff['removed'])} removed")

            if update_db:
                db_by_sku = {item.sku: item for item in db_items}
                updated = sum(
                    1 for sku, info in by_sku.items()
                    if info["price"] is not None and sku in db_by_sku
                    and not setattr(db_by_sku[sku], "msrp", info["price"])  # side-effect
                )
                db.session.commit()
                log(f"Updated {updated} MSRP prices in database")

                hits = check_wishlist_targets()
                if hits:
                    log(f"Wishlist targets met: {len(hits)}")
                    if DISCORD_WEBHOOK_URL:
                        lines = ["**🎯 Cutco Wishlist — Price Targets Met**"]
                        for hit in hits:
                            lines.append(
                                f"• **{hit['person']}** — {hit['item']} (#{hit['sku']}): "
                                f"MSRP ${hit['msrp']:.2f} ≤ target ${hit['target']:.2f}"
                            )
                        _notify_discord("\n".join(lines))

            job = _read_msrp_job()
            job.update({"status": "done", "results": diff,
                        "finished_at": date.today().isoformat()})
            _write_msrp_job(job)

    except Exception as exc:
        logger.error("MSRP diff job failed: %s", exc)
        job = _read_msrp_job()
        job.update({"status": "error", "error": str(exc),
                    "finished_at": date.today().isoformat()})
        _write_msrp_job(job)


def _extract_sku_from_href(href: str) -> str | None:
    """Pull a base SKU from a /p/ product URL.

    Strips accessory and color suffixes so that bundle SKUs (e.g. 4135CSH —
    the 4-inch vegetable knife with sheath) resolve to the base item SKU
    (4135) rather than being treated as a separate catalog entry.

    Suffix stripping order:
      1. Remove trailing 'SH' (sheath bundle indicator)
      2. Remove trailing color letter (C / W / R / B)

    Returns None if the URL is not a product link.
    """
    parts = href.rstrip("/").split("/")
    slug = parts[-1].split("?")[0].split("&")[0].upper()
    if not slug:
        return None
    # If slug starts with digits, extract the leading numeric+letter portion.
    # This handles "1720-PETITE-CHEF" → "1720" as well as "4135CSH" → "4135C".
    lead = re.match(r'^(\d{3,}[A-Z]{0,3})', slug)
    if lead:
        candidate = lead.group(1)
    elif any(char.isdigit() for char in slug) and len(slug) <= 12:
        # Short slug with embedded digits (e.g. "ABC1234")
        candidate = slug
    else:
        return None
    # Strip sheath suffix, then all trailing color letters, to get the base SKU
    if candidate.endswith("SH"):
        candidate = candidate[:-2]
    stripped = re.sub(r"[A-Z]+$", "", candidate)
    if stripped and stripped.isdigit() and len(stripped) >= 2:
        candidate = stripped
    return candidate or None


def _discover_categories() -> list[tuple[str, str]]:
    """Scrape the Cutco shop index to discover all category pages automatically.

    Returns a list of (category_name, url) tuples found in the shop navigation.
    Falls back to an empty list if the shop page cannot be reached.
    """
    # Try the homepage — it reliably has nav links to all /shop/ categories.
    # The /shop/ index itself 404s on cutco.com.
    discovery_urls = [
        "https://www.cutco.com/",
        "https://www.cutco.com/products/knives/",
    ]
    discovered = []
    seen_slugs  = set()
    try:
        resp = None
        for url in discovery_urls:
            try:
                resp = requests.get(url, headers=SCRAPE_HEADERS, timeout=REQUEST_TIMEOUT)
                if resp.status_code == 200:
                    break
            except Exception:
                continue
        if resp is None or resp.status_code != 200:
            raise RuntimeError("No discovery URL returned 200")
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "html.parser")
        for anchor in soup.select("a[href*='/shop/']"):
            href = anchor.get("href", "").rstrip("/")
            slug = href.split("/shop/")[-1].split("?")[0]
            # Skip empty slugs, set pages, and already-seen slugs
            if not slug or slug in seen_slugs or "knife-set" in slug or "set" == slug:
                continue
            url  = href if href.startswith("http") else f"https://www.cutco.com{href}"
            # Derive a human-readable name: title-case the slug
            name = slug.replace("-", " ").title()
            seen_slugs.add(slug)
            discovered.append((name, url))
        logger.info("Discovered %d categories from shop index", len(discovered))
    except Exception as exc:
        logger.warning("Category discovery failed: %s", exc)
    return discovered


def _build_category_list() -> list[tuple[str, str]]:
    """Merge auto-discovered categories with the hardcoded SCRAPE_CATEGORIES list.

    Hardcoded entries take precedence — their names override auto-discovered
    names for the same URL slug, and they are always included even if discovery
    fails.  Discovered categories not already in SCRAPE_CATEGORIES are appended.
    """
    # Build a slug→(name, url) map from hardcoded list (authoritative)
    def slug_of(url):
        return url.rstrip("/").split("/shop/")[-1].split("?")[0].lower()

    known = {slug_of(url): (name, url) for name, url in SCRAPE_CATEGORIES}

    # Add discovered categories that aren't already covered
    for name, url in _discover_categories():
        slug = slug_of(url)
        if slug not in known:
            known[slug] = (name, url)

    return list(known.values())


def _fetch_sku_from_page(url: str) -> tuple[str | None, str | None]:
    """Fetch a product page and return (sku, name) from on-page content.

    Used for pure-slug URLs like /p/fishermans-solution where no numeric SKU
    appears in the URL itself.  Tries several extraction strategies in order:
      1. JSON-LD structured data (<script type="application/ld+json">)
      2. Cutco JS page variables: prPageId / defaultWebItemSingle
      3. Generic inline JS — "sku":"1769" in raw HTML
      4. Meta tag (product:retailer_item_id or similar)
      5. On-page visible text matching "#XXXX" (scripts/styles stripped first)
      6. Broader keyword context (Model/Item/SKU followed by a number)
    """
    try:
        resp = requests.get(url, headers=SCRAPE_HEADERS, timeout=REQUEST_TIMEOUT)
        if resp.status_code != 200:
            logger.info("SKU fetch: HTTP %d for %s", resp.status_code, url)
            return None, None
        raw_html = resp.text
        soup = BeautifulSoup(raw_html, "html.parser")
        sku = None
        strategy_log: list[str] = []

        # Strategy 0: SKU embedded in the URL slug (e.g. /p/677-super-shears-gift-box).
        # Most unambiguous source — if the URL carries a numeric prefix we trust it.
        url_slug_sku = _extract_sku_from_href(url)
        if url_slug_sku:
            sku = url_slug_sku
            strategy_log.append(f"slug={sku}")

        # Strategy 1: JSON-LD structured data.
        if not sku:
            for ld in soup.find_all("script", type="application/ld+json"):
                try:
                    data = json.loads(ld.string or "")
                    entries = data if isinstance(data, list) else [data]
                    for entry in entries:
                        if isinstance(entry, dict) and entry.get("@type") == "Product":
                            sku_val = entry.get("sku") or entry.get("productID")
                            if sku_val:
                                sku = str(sku_val).strip().upper()
                                break
                except (json.JSONDecodeError, AttributeError):
                    pass
                if sku:
                    break
            if sku:
                strategy_log.append(f"json-ld={sku}")

        # Strategy 2: Cutco-specific JS page variables.
        # prPageId is reliable for standalone products but on gift box pages
        # it may reference the contained knife's SKU rather than the gift box.
        #   const prPageId = "677";
        #   const defaultWebItemSingle = "1886BK";
        if not sku:
            for js_var_name in ("prPageId", "defaultWebItemSingle"):
                sku_match = re.search(
                    rf"""(?:const|var|let)\s+{js_var_name}\s*=\s*["']([^"']+)["']""",
                    raw_html)
                if sku_match:
                    # Capture leading digits plus optional -N suffix so sheath
                    # SKUs like "4135-2" are preserved distinct from the knife "4135".
                    digits = re.match(r'^(\d{2,}(?:-\d+)?)', sku_match.group(1).strip())
                    if digits:
                        sku = digits.group(1)
                        break
            if sku:
                strategy_log.append(f"prPageId={sku}")

        # Strategy 3: generic inline JS — catches patterns like:
        #   "sku":"1769"  |  'sku': '1769'  |  sku: 1769
        if not sku:
            sku_match = re.search(
                r"""["']?sku["']?\s*:\s*["']?(\d{2,4}[A-Z]{0,2})["']?""",
                raw_html, re.IGNORECASE)
            if sku_match:
                sku = sku_match.group(1).upper()
                strategy_log.append(f"inline-js={sku}")

        # Strategy 4: meta tags (Open Graph / Schema product SKU)
        if not sku:
            for attr in ("product:retailer_item_id", "product:sku"):
                tag = soup.find("meta", property=attr) or soup.find("meta", attrs={"name": attr})
                if tag and tag.get("content", "").strip():
                    sku = tag["content"].strip().upper()
                    strategy_log.append(f"meta={sku}")
                    break

        # Remove script/style blocks before text search so CSS hex colors
        # (e.g. brand blue #0073A4) don't get matched as SKUs.
        for noise in soup.find_all(["script", "style"]):
            noise.decompose()

        # Strategy 5: on-page visible text containing "#XXXX".
        if not sku:
            sku_text = soup.find(string=re.compile(r"#\d{2,4}[A-Z]?\b"))
            if sku_text:
                sku_match = re.search(r"#(\d{2,4}[A-Z]?)\b", sku_text.strip(), re.IGNORECASE)
                if sku_match:
                    sku = sku_match.group(1).upper()
                    strategy_log.append(f"visible-text={sku}")

        # Strategy 6: keyword context — "Model 1769", "Item No. 1769", etc.
        if not sku:
            page_text = soup.get_text(" ", strip=True)
            sku_match = re.search(
                r"(?:model|item|sku|product)\s*(?:no\.?|number|#)?\s*[:#]?\s*(\d{2,4}[A-Z]?)\b",
                page_text, re.IGNORECASE)
            if sku_match:
                sku = sku_match.group(1).upper()
                strategy_log.append(f"keyword={sku}")

        # Normalise: strip all trailing color/variant letters to get the base SKU
        stripped = re.sub(r"[A-Z]+$", "", sku or "")
        if stripped and stripped.isdigit() and len(stripped) >= 2:
            sku = stripped

        # Reject CSS hex colors — 6 hex chars like "0073A4" are never a SKU
        if sku and re.fullmatch(r"[0-9A-F]{6}", sku, re.IGNORECASE):
            sku = None

        page_heading = soup.find("h1")
        name = page_heading.get_text(strip=True) if page_heading else None
        logger.info("SKU fetch: %s → sku=%s [%s] name=%s",
                    url, sku, ", ".join(strategy_log) or "none", name)
        return sku, name
    except Exception as exc:
        logger.warning("SKU fetch failed: %s — %s", url, exc)
        return None, None


def scrape_catalog() -> tuple[list[dict], list[tuple[str, str]]]:
    """Scrape all item categories.

    Returns (items, set_candidates) where set_candidates is a list of
    (name, url) tuples for products filtered out as bundles/sets so that
    scrape_sets() can visit their detail pages for member-SKU extraction.
    """
    results        = []
    set_candidates: list[tuple[str, str]] = []
    seen_skus      = set()
    seen_set_urls: set[str] = set()
    # pure-slug items queued for a single parallel fetch pass at the end:
    # list of (prod_url, cat_name, name_from_category_page)
    slug_queue: list[tuple[str, str, str | None]] = []
    seen_slug_urls: set[str] = set()

    categories = _build_category_list()
    logger.info("Scraping %d categories", len(categories))
    for cat_name, cat_url in categories:
        if cat_name in SYNC_BLOCKED_CATEGORIES:
            continue
        try:
            resp = requests.get(cat_url, headers=SCRAPE_HEADERS, timeout=REQUEST_TIMEOUT)
            resp.raise_for_status()
            soup = BeautifulSoup(resp.text, "html.parser")

            # Walk the DOM in order, collecting product links and stopping
            # once a bundle/gift section heading is encountered.
            # Only stop at sections that explicitly indicate gift bundles or
            # additional/accessory groupings — not generic headings like "set"
            # or "collection" which appear in normal product sections too.
            product_links = []
            for element in soup.descendants:
                if element.name in ("h2", "h3", "h4") and product_links:
                    if any(kw in element.get_text(strip=True).lower()
                           for kw in _BUNDLE_KEYWORDS):
                        logger.debug("Bundle section detected on %s: '%s'",
                                     cat_url, element.get_text(strip=True))
                        break
                if element.name == "a" and "/p/" in element.get("href", ""):
                    product_links.append(element)

            logger.debug("%s: found %d /p/ links — %s",
                         cat_name, len(product_links),
                         [a.get("href", "") for a in product_links[:10]])

            # Deduplicate: strip &view=product variants so each product is
            # processed at most once per category, but keep the full href
            # (including &view=product) for the actual fetch so the server
            # returns the full product page with JSON-LD.
            seen_hrefs: set[str] = set()
            unique_links = []
            for anchor in product_links:
                full_href = anchor.get("href", "")
                base_href = full_href.split("&")[0]
                if base_href not in seen_hrefs:
                    seen_hrefs.add(base_href)
                    unique_links.append((anchor, full_href))

            for anchor, href in unique_links:
                base_href = href.split("&")[0]
                sku = _extract_sku_from_href(base_href)
                prod_url = href if href.startswith("http") else f"https://www.cutco.com{href}"

                name_el = anchor.find(["h2", "h3"])
                if not name_el and anchor.parent:
                    name_el = anchor.parent.find(["h2", "h3"])
                name = name_el.get_text(strip=True) if name_el else None

                # Sheaths: URL-based extraction returns the parent knife's
                # SKU (e.g. /p/4135-2 → 4135), not the sheath's own model number.
                # Force all sheath URLs through page fetch so prPageId gives the
                # real sheath SKU (which includes the -2 suffix).
                # Apply this even when sheaths appear in other categories (e.g. Storage).
                if cat_name == "Sheaths" or (name and "sheath" in name.lower() and "with sheath" not in name.lower()):
                    sku = None

                if not sku:
                    # Pure-slug URL — queue for parallel page fetch after this loop.
                    # Always fetch with &view=product so the server returns the full
                    # product page with the prPageId JS variable.
                    if "&view=product" not in prod_url:
                        prod_url = prod_url + "&view=product"
                    if prod_url not in seen_slug_urls:
                        seen_slug_urls.add(prod_url)
                        slug_queue.append((prod_url, cat_name, name))
                    continue

                if sku in seen_skus or not name:
                    continue
                if _is_set_product(name):
                    if prod_url not in seen_set_urls:
                        seen_set_urls.add(prod_url)
                        set_candidates.append((name, prod_url))
                    continue
                seen_skus.add(sku)
                results.append(dict(name=name, sku=sku,
                                    category=_resolve_category(sku, cat_name, name),
                                    url=prod_url))
            time.sleep(0.4)
        except Exception as exc:
            logger.warning("Scrape failed for %s: %s", cat_url, exc)

    # Parallel fetch for all pure-slug product pages collected above.
    # Using a thread pool avoids serialising hundreds of HTTP round-trips.
    if slug_queue:
        logger.info("Fetching %d pure-slug product pages (parallel)", len(slug_queue))
        added_from_slugs = 0
        with ThreadPoolExecutor(max_workers=6) as pool:
            future_map = {
                pool.submit(_fetch_sku_from_page, prod_url): (prod_url, cat_name, cat_name_hint)
                for prod_url, cat_name, cat_name_hint in slug_queue
            }
            for future in as_completed(future_map):
                prod_url, cat_name, cat_page_name = future_map[future]
                sku, page_name = future.result()
                name = cat_page_name or page_name
                if not sku or sku in seen_skus or not name:
                    continue
                if _is_set_product(name):
                    if prod_url not in seen_set_urls:
                        seen_set_urls.add(prod_url)
                        set_candidates.append((name, prod_url))
                    continue
                seen_skus.add(sku)
                results.append(dict(name=name, sku=sku,
                                    category=_resolve_category(sku, cat_name, name),
                                    url=prod_url))
                added_from_slugs += 1
        logger.info("Slug queue: %d pages fetched, %d items added", len(slug_queue), added_from_slugs)

    return results, set_candidates


def scrape_sets(
    extra_candidates: list[tuple[str, str]] | None = None,
) -> list[dict]:
    """
    Scrape the knife-sets listing page, then visit each set detail page to
    extract member item SKUs from the Set Pieces section image URLs.

    extra_candidates is an optional list of (name, url) tuples for bundle/set
    products discovered during the catalog scrape (e.g. gift boxes, combos)
    that do not appear on the knife-sets listing page.

    Returns a list of dicts:
      { name, sku, url, member_skus: [str, ...] }
    """
    results = []
    seen_slugs: set[str] = set()
    set_links = []

    try:
        resp = requests.get(SCRAPE_SETS_URL, headers=SCRAPE_HEADERS, timeout=REQUEST_TIMEOUT)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "html.parser")
        for anchor in soup.select("a[href*='/p/']"):
            href = anchor.get("href", "")
            if not href:
                continue
            name_el = anchor.find(["h2", "h3"])
            if not name_el and anchor.parent:
                name_el = anchor.parent.find(["h2", "h3"])
            name = name_el.get_text(strip=True) if name_el else None
            if not name:
                continue
            url = href if href.startswith("http") else f"https://www.cutco.com{href}"
            slug = href.rstrip("/").split("/")[-1].split("&")[0]
            if slug not in seen_slugs:
                seen_slugs.add(slug)
                set_links.append(dict(name=name, slug=slug, url=url))
    except Exception as exc:
        logger.warning("Scrape failed for sets listing: %s", exc)

    # Merge in gift sets / bundles found during the catalog scrape
    for name, url in (extra_candidates or []):
        slug = url.rstrip("/").split("/")[-1].split("&")[0]
        if slug not in seen_slugs:
            seen_slugs.add(slug)
            set_links.append(dict(name=name, slug=slug, url=url))

    # Fetch each set detail page in parallel to extract SKU and member items.
    # Image URLs look like: /products/rolo/1720C-h.jpg — SKU is the numeric prefix.
    sku_pattern = re.compile(r"/rolo/([0-9]+[A-Z]?)-h\.", re.IGNORECASE)

    def _fetch_set_detail(set_link: dict) -> dict | None:
        fetch_url = set_link["url"]
        if "&view=product" not in fetch_url:
            fetch_url += "&view=product"
        try:
            set_sku, _ = _fetch_sku_from_page(fetch_url)
            resp = requests.get(fetch_url, headers=SCRAPE_HEADERS, timeout=REQUEST_TIMEOUT)
            resp.raise_for_status()
            detail = BeautifulSoup(resp.text, "html.parser")
            member_skus = []
            seen_member: set[str] = set()
            for img in detail.select("img[src*='/rolo/']"):
                match = sku_pattern.search(img.get("src", ""))
                if match:
                    raw = match.group(1).upper()
                    base_sku = re.sub(r"[A-Z]+$", "", raw) if len(raw) > 2 else raw
                    # Skip year-like numbers (2000–2099) — these appear in
                    # marketing/packaging image filenames, not product SKUs.
                    if re.fullmatch(r"20\d{2}", base_sku):
                        continue
                    if base_sku not in seen_member:
                        seen_member.add(base_sku)
                        member_skus.append(base_sku)
            logger.debug("Set '%s': sku=%s, %d members", set_link["name"], set_sku, len(member_skus))
            return dict(name=set_link["name"], sku=set_sku, url=set_link["url"], member_skus=member_skus)
        except Exception as exc:
            logger.warning("Scrape failed for set %s: %s", set_link["url"], exc)
            return None

    logger.info("Fetching %d set detail pages (parallel)", len(set_links))
    with ThreadPoolExecutor(max_workers=6) as pool:
        for result in pool.map(_fetch_set_detail, set_links):
            if result is not None:
                results.append(result)

    logger.info("Sets scraped: %d", len(results))
    return results


def get_or_create_set(name: str) -> "Set":
    """Return existing Set by name or create a new one."""
    item_set = Set.query.filter(db.func.lower(Set.name) == name.lower()).first()
    if not item_set:
        item_set = Set(name=name)
        db.session.add(item_set)
        db.session.flush()
    return item_set

# ── Template context ──────────────────────────────────────────────────────────

@app.context_processor
def inject_globals():
    return dict(app_version=APP_VERSION, is_admin=is_admin)

# ── Health & Version ──────────────────────────────────────────────────────────

@app.route("/health")
def health():
    try:
        db.session.execute(db.text("SELECT 1"))
        return jsonify(status="ok", version=APP_VERSION), 200
    except Exception as exc:
        logger.error("Health check failed: %s", exc)
        return jsonify(status="error", detail=str(exc)), 500


@app.route("/version")
def version():
    return jsonify(version=APP_VERSION)

# ── Dashboard ─────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    stats = dict(
        item_count = Item.query.count(),
        unicorns = Item.query.filter(db.or_(
                       Item.is_unicorn,
                       Item.variants.any(ItemVariant.is_unicorn == True)  # noqa: E712
                   )).count(),
        people   = Person.query.count(),
        owned    = Ownership.query.filter_by(status="Owned").count(),
        wishlist = Ownership.query.filter_by(status="Wishlist").count(),
        variants = ItemVariant.query.filter(ItemVariant.color != UNKNOWN_COLOR).count(),
        sets     = Set.query.count(),
    )
    people = Person.query.order_by(Person.name).all()
    recent = Ownership.query.order_by(Ownership.id.desc()).limit(10).all()
    return render_template("index.html", stats=stats, people=people, recent=recent)

# ── Catalog ───────────────────────────────────────────────────────────────────

@app.route("/catalog")
def catalog():
    search_query = request.args.get("q", "").strip()
    cat_filter = request.args.get("category", "")
    unicorn_f  = request.args.get("unicorn", "")
    sort       = request.args.get("sort", "name")
    direction  = request.args.get("dir", "asc")

    query = Item.query
    if search_query:
        query = query.filter(
            db.or_(Item.name.ilike(f"%{search_query}%"), Item.sku.ilike(f"%{search_query}%")))
    if cat_filter:
        query = query.filter(Item.category == cat_filter)
    if unicorn_f == "1":
        query = query.filter(db.or_(
            Item.is_unicorn,
            Item.variants.any(ItemVariant.is_unicorn == True)  # noqa: E712
        ))

    col   = getattr(Item, sort, Item.name)
    items = (query
             .options(selectinload(Item.variants), selectinload(Item.sets))
             .order_by(col.desc() if direction == "desc" else col)
             .all())

    categories = [row[0] for row in
                  db.session.query(Item.category)
                  .filter(Item.category.isnot(None))
                  .distinct().order_by(Item.category).all()]

    return render_template("catalog.html", items=items, categories=categories,
                           q=search_query, cat_filter=cat_filter, unicorn_f=unicorn_f,
                           sort=sort, direction=direction,
                           edge_types=EDGE_TYPES,
                           UNKNOWN_COLOR=UNKNOWN_COLOR)


@app.route("/catalog/add", methods=["GET", "POST"])
def catalog_add():
    if request.method == "POST":
        item = Item(
            name       = request.form["name"].strip(),
            sku        = request.form.get("sku", "").strip().upper() or None,
            category   = request.form.get("category", "").strip() or None,
            edge_type  = request.form.get("edge_type", "Unknown"),
            is_unicorn = request.form.get("is_unicorn") == "on",
            in_catalog = request.form.get("in_catalog") == "on",
            cutco_url  = request.form.get("cutco_url", "").strip() or None,
            notes      = request.form.get("notes", "").strip() or None,
        )
        db.session.add(item)
        db.session.flush()
        ensure_unknown_variant(item)
        colors = [c.strip() for c in request.form.get("colors", "").split(",") if c.strip()]
        for color in colors:
            if color != UNKNOWN_COLOR:
                db.session.add(ItemVariant(item_id=item.id, color=color))
        db.session.commit()
        logger.info("Item added: %s (SKU: %s)", item.name, item.sku or "none")
        flash(f'Added "{item.name}" to catalog.', "success")
        return redirect(url_for("catalog"))

    return render_template("item_form.html", item=None,
                           edge_types=EDGE_TYPES, action="Add",
                           UNKNOWN_COLOR=UNKNOWN_COLOR,
                           all_sets=Set.query.order_by(Set.name).all())


@app.route("/catalog/<int:iid>/edit", methods=["GET", "POST"])
def catalog_edit(iid):
    item = Item.query.get_or_404(iid)
    if request.method == "POST":
        item.name       = request.form["name"].strip()
        item.sku        = request.form.get("sku", "").strip().upper() or None
        item.category   = request.form.get("category", "").strip() or None
        item.edge_type  = request.form.get("edge_type", "Unknown")
        item.is_unicorn = request.form.get("is_unicorn") == "on"
        item.in_catalog = request.form.get("in_catalog") == "on"
        item.cutco_url  = request.form.get("cutco_url", "").strip() or None
        item.notes      = request.form.get("notes", "").strip() or None

        # Update set memberships
        selected_set_ids = set(int(set_id_str) for set_id_str in request.form.getlist("set_ids"))
        item.sets = Set.query.filter(Set.id.in_(selected_set_ids)).all()

        db.session.commit()
        logger.info("Item updated: %s (SKU: %s)", item.name, item.sku or "none")
        flash(f'Updated "{item.name}".', "success")
        return redirect(url_for("catalog"))

    return render_template("item_form.html", item=item,
                           edge_types=EDGE_TYPES, action="Edit",
                           UNKNOWN_COLOR=UNKNOWN_COLOR,
                           all_sets=Set.query.order_by(Set.name).all())


@app.route("/catalog/<int:iid>/delete", methods=["POST"])
def catalog_delete(iid):
    item = Item.query.get_or_404(iid)
    name = item.name
    db.session.delete(item)
    db.session.commit()
    logger.info("Item deleted: %s", name)
    flash(f'Deleted "{name}".', "info")
    return redirect(url_for("catalog"))

# ── Variants ──────────────────────────────────────────────────────────────────

@app.route("/catalog/<int:iid>/variants")
def variants(iid):
    item = Item.query.get_or_404(iid)
    return render_template("variants.html", item=item, UNKNOWN_COLOR=UNKNOWN_COLOR)


@app.route("/catalog/<int:iid>/variants/add", methods=["POST"])
def variant_add(iid):
    item = Item.query.get_or_404(iid)
    color = request.form.get("color", "").strip()
    if not color:
        flash("Color is required.", "error")
        return redirect(url_for("variants", iid=iid))
    if any(v.color.lower() == color.lower() for v in item.variants):
        flash(f'"{color}" already exists for this item.', "error")
        return redirect(url_for("variants", iid=iid))
    db.session.add(ItemVariant(item_id=iid, color=color,
                               notes=request.form.get("notes", "").strip() or None))
    db.session.commit()
    logger.info("Variant added: %s → %s", item.name, color)
    flash(f'Added variant "{color}".', "success")
    return redirect(url_for("variants", iid=iid))


@app.route("/variants/<int:vid>/edit", methods=["POST"])
def variant_edit(vid):
    variant = ItemVariant.query.get_or_404(vid)
    iid     = variant.item_id
    color   = request.form.get("color", "").strip()
    if not color:
        flash("Color cannot be empty.", "error")
        return redirect(url_for("variants", iid=iid))
    variant.color      = color
    variant.notes      = request.form.get("notes", "").strip() or None
    variant.is_unicorn = request.form.get("is_unicorn") == "on"
    db.session.commit()
    logger.info("Variant updated: item %d → %s", iid, color)
    flash(f'Updated to "{color}".', "success")
    return redirect(url_for("variants", iid=iid))


@app.route("/variants/<int:vid>/delete", methods=["POST"])
def variant_delete(vid):
    variant = ItemVariant.query.get_or_404(vid)
    if len(variant.item.variants) == 1:
        flash("Cannot delete the only variant. Add another first.", "error")
        return redirect(url_for("variants", iid=variant.item_id))
    iid = variant.item_id
    db.session.delete(variant)
    db.session.commit()
    logger.info("Variant deleted: item %d", iid)
    flash("Variant removed.", "info")
    return redirect(url_for("variants", iid=iid))

# ── Sets ──────────────────────────────────────────────────────────────────────

@app.route("/sets")
def sets_list():
    all_sets = Set.query.order_by(Set.name).all()
    return render_template("sets.html", sets=all_sets)


@app.route("/sets/add", methods=["GET", "POST"])
def set_add():
    if request.method == "POST":
        name = request.form["name"].strip()
        if Set.query.filter(db.func.lower(Set.name) == name.lower()).first():
            flash(f'Set "{name}" already exists.', "error")
            return redirect(url_for("set_add"))
        item_set = Set(name=name, notes=request.form.get("notes", "").strip() or None)
        db.session.add(item_set)
        db.session.commit()
        logger.info("Set created: %s", name)
        flash(f'Created set "{name}".', "success")
        return redirect(url_for("sets_list"))
    return render_template("set_form.html", set=None, action="Add")


@app.route("/sets/<int:sid>/edit", methods=["GET", "POST"])
def set_edit(sid):
    item_set = Set.query.get_or_404(sid)
    if request.method == "POST":
        item_set.name  = request.form["name"].strip()
        item_set.notes = request.form.get("notes", "").strip() or None
        db.session.commit()
        logger.info("Set updated: %s", item_set.name)
        flash(f'Updated set "{item_set.name}".', "success")
        return redirect(url_for("sets_list"))
    return render_template("set_form.html", set=item_set, action="Edit")


@app.route("/sets/<int:sid>/delete", methods=["POST"])
def set_delete(sid):
    item_set = Set.query.get_or_404(sid)
    name = item_set.name
    db.session.delete(item_set)
    db.session.commit()
    logger.info("Set deleted: %s", name)
    flash(f'Deleted set "{name}".', "info")
    return redirect(url_for("sets_list"))


@app.route("/sets/<int:sid>")
def set_detail(sid):
    item_set = Set.query.get_or_404(sid)
    return render_template("set_detail.html", set=item_set, UNKNOWN_COLOR=UNKNOWN_COLOR)

# ── Catalog Sync ──────────────────────────────────────────────────────────────

@app.route("/catalog/sync")
def catalog_sync():
    if not is_admin():
        flash("Admin access required.", "error")
        return redirect(url_for("catalog"))

    scraped, set_candidates = scrape_catalog()
    existing_skus = {item.sku for item in Item.query.filter(Item.sku.isnot(None)).all()}
    new_items = [i for i in scraped if i["sku"] not in existing_skus]

    from collections import OrderedDict
    _grouped_unsorted: dict = {}
    for item in new_items:
        _grouped_unsorted.setdefault(item["category"], []).append(item)
    def _sku_sort_key(item):
        sku = item.get("sku") or ""
        sku_num_match = re.match(r"(\d+)", sku)
        return (0, int(sku_num_match.group(1)), sku) if sku_num_match else (1, 0, sku)

    grouped = OrderedDict(
        (cat, sorted(items, key=_sku_sort_key))
        for cat, items in sorted(_grouped_unsorted.items(), key=lambda cat_items_pair: cat_items_pair[0].lower())
    )

    # Also scrape sets for preview — pass gift/bundle candidates from catalog scrape
    scraped_sets  = scrape_sets(extra_candidates=set_candidates)
    existing_sets = {existing_set.name.lower() for existing_set in Set.query.all()}
    new_sets      = sorted(
        (s for s in scraped_sets if s["name"].lower() not in existing_sets),
        key=_sku_sort_key,
    )

    logger.info("Sync: %d items scraped, %d new; %d sets scraped, %d new",
                len(scraped), len(new_items), len(scraped_sets), len(new_sets))

    return render_template("sync_preview.html",
                           grouped=grouped,
                           new_items=new_items,
                           scraped_total=len(scraped),
                           new_sets=new_sets,
                           scraped_sets_total=len(scraped_sets),
                           blocked_categories=sorted(SYNC_BLOCKED_CATEGORIES))


@app.route("/catalog/sync/confirm", methods=["POST"])
def catalog_sync_confirm():
    if not is_admin():
        flash("Admin access required.", "error")
        return redirect(url_for("catalog"))

    # ── Items ──────────────────────────────────────────────────────────────
    selected = set(request.form.getlist("selected_skus"))
    item_data = {}
    for key, val in request.form.items():
        for prefix in ("name_", "category_", "url_"):
            if key.startswith(prefix):
                sku = key[len(prefix):]
                item_data.setdefault(sku, {})[prefix.rstrip("_")] = val

    added_items = 0
    for sku in selected:
        if Item.query.filter_by(sku=sku).first():
            continue
        data = item_data.get(sku, {})
        item = Item(name=data.get("name", sku), sku=sku,
                    category=data.get("category"), cutco_url=data.get("url"),
                    in_catalog=True, is_unicorn=False, edge_type="Unknown")
        db.session.add(item)
        db.session.flush()
        ensure_unknown_variant(item)
        added_items += 1

    db.session.flush()  # ensure all new items have IDs before set linkage

    # ── Sets ───────────────────────────────────────────────────────────────
    selected_sets = set(request.form.getlist("selected_sets"))
    added_sets    = 0
    linked_items  = 0

    # Build fresh SKU→Item map including items just added above
    sku_to_item = {item.sku.upper(): item for item in Item.query.filter(Item.sku.isnot(None)).all()}

    set_count = int(request.form.get("set_count", 0))
    for i in range(set_count):
        set_name = request.form.get(f"set_name_{i}", "").strip()
        if not set_name or set_name not in selected_sets:
            continue
        member_skus = [raw.strip() for raw in
                       request.form.get(f"set_members_{i}", "").split("|") if raw.strip()]
        set_sku = request.form.get(f"set_sku_{i}", "").strip() or None

        item_set = get_or_create_set(set_name)
        if item_set.id is None:
            added_sets += 1
        if set_sku and not item_set.sku:
            item_set.sku = set_sku

        for msku in member_skus:
            item = sku_to_item.get(msku.upper())
            if item and item not in item_set.items:
                item_set.items.append(item)
                linked_items += 1

    db.session.commit()

    parts = []
    if added_items:
        parts.append(f"{added_items} item{'s' if added_items != 1 else ''}")
    if added_sets:
        parts.append(f"{added_sets} set{'s' if added_sets != 1 else ''}")
    if linked_items:
        parts.append(f"{linked_items} set membership{'s' if linked_items != 1 else ''}")
    flash("Sync complete — added " + (", ".join(parts) if parts else "nothing new") + ".", "success")
    return redirect(url_for("catalog"))

# ── People ────────────────────────────────────────────────────────────────────

@app.route("/people")
def people():
    persons = Person.query.order_by(Person.name).all()
    counts  = {person.id: Ownership.query.filter_by(person_id=person.id, status="Owned").count()
               for person in persons}
    return render_template("people.html", persons=persons, counts=counts)


@app.route("/people/add", methods=["GET", "POST"])
def people_add():
    if request.method == "POST":
        person = Person(name=request.form["name"].strip(),
                        notes=request.form.get("notes", "").strip() or None)
        db.session.add(person)
        db.session.commit()
        logger.info("Person added: %s", person.name)
        flash(f"Added {person.name}.", "success")
        return redirect(url_for("people"))
    return render_template("person_form.html", person=None, action="Add")


@app.route("/people/<int:pid>/edit", methods=["GET", "POST"])
def people_edit(pid):
    person = Person.query.get_or_404(pid)
    if request.method == "POST":
        person.name  = request.form["name"].strip()
        person.notes = request.form.get("notes", "").strip() or None
        db.session.commit()
        logger.info("Person updated: %s", person.name)
        flash(f"Updated {person.name}.", "success")
        return redirect(url_for("people"))
    return render_template("person_form.html", person=person, action="Edit")


@app.route("/people/<int:pid>/delete", methods=["POST"])
def people_delete(pid):
    person = Person.query.get_or_404(pid)
    name   = person.name
    db.session.delete(person)
    db.session.commit()
    logger.info("Person deleted: %s", name)
    flash(f"Removed {name}.", "info")
    return redirect(url_for("people"))


@app.route("/people/<int:pid>/collection")
def person_collection(pid):
    person     = Person.query.get_or_404(pid)
    ownerships = (Ownership.query.filter_by(person_id=pid)
                  .order_by(Ownership.status).all())

    owned_item_ids = {o.variant.item_id for o in ownerships if o.status == "Owned"}
    all_items      = Item.query.order_by(Item.name).all()
    item_gaps      = [item for item in all_items if item.id not in owned_item_ids]

    variant_gaps = []
    for item in all_items:
        real_variants = [variant for variant in item.variants if variant.color != UNKNOWN_COLOR]
        if not real_variants:
            continue
        owned_variant_ids = {ownership.variant_id for ownership in ownerships
                             if ownership.variant.item_id == item.id and ownership.status == "Owned"}
        missing = [variant for variant in real_variants if variant.id not in owned_variant_ids]
        if missing:
            variant_gaps.append((item, missing))

    return render_template("collection.html", person=person,
                           ownerships=ownerships,
                           item_gaps=item_gaps,
                           variant_gaps=variant_gaps,
                           status_options=STATUS_OPTIONS,
                           UNKNOWN_COLOR=UNKNOWN_COLOR)

# ── Ownership CRUD ────────────────────────────────────────────────────────────

@app.route("/ownership/add", methods=["GET", "POST"])
def ownership_add():
    person_id  = request.args.get("person_id", type=int)
    item_id    = request.args.get("item_id", type=int)
    variant_id = request.args.get("variant_id", type=int)

    if request.method == "POST":
        person_id  = int(request.form["person_id"])
        variant_id = int(request.form["variant_id"])
        if Ownership.query.filter_by(person_id=person_id, variant_id=variant_id).first():
            flash("That person already has an entry for that variant.", "error")
            return redirect(url_for("person_collection", pid=person_id))
        raw_target = request.form.get("target_price", "").strip()
        try:
            target_price = float(raw_target) if raw_target else None
        except ValueError:
            target_price = None
        db.session.add(Ownership(
            person_id    = person_id,
            variant_id   = variant_id,
            status       = request.form.get("status", "Owned"),
            target_price = target_price,
            notes        = request.form.get("notes", "").strip() or None,
        ))
        db.session.commit()
        logger.info("Ownership added: person %d, variant %d", person_id, variant_id)
        flash("Entry logged.", "success")
        return redirect(url_for("person_collection", pid=person_id))

    sel_item = Item.query.get(item_id) if item_id else None
    return render_template("ownership_form.html", ownership=None,
                           people_list=Person.query.order_by(Person.name).all(),
                           items_list=Item.query.order_by(Item.name).all(),
                           status_options=STATUS_OPTIONS,
                           sel_person_id=person_id,
                           sel_item_id=item_id,
                           sel_variant_id=variant_id,
                           sel_item=sel_item,
                           action="Add",
                           UNKNOWN_COLOR=UNKNOWN_COLOR)


@app.route("/ownership/<int:oid>/edit", methods=["GET", "POST"])
def ownership_edit(oid):
    ownership = Ownership.query.get_or_404(oid)
    if request.method == "POST":
        ownership.status = request.form.get("status", "Owned")
        raw_target = request.form.get("target_price", "").strip()
        try:
            ownership.target_price = float(raw_target) if raw_target else None
        except ValueError:
            ownership.target_price = None
        ownership.notes  = request.form.get("notes", "").strip() or None
        db.session.commit()
        logger.info("Ownership updated: id %d → %s", oid, ownership.status)
        flash("Updated.", "success")
        return redirect(url_for("person_collection", pid=ownership.person_id))

    return render_template("ownership_form.html", ownership=ownership,
                           people_list=Person.query.order_by(Person.name).all(),
                           items_list=Item.query.order_by(Item.name).all(),
                           status_options=STATUS_OPTIONS,
                           sel_person_id=ownership.person_id,
                           sel_item_id=ownership.variant.item_id,
                           sel_variant_id=ownership.variant_id,
                           sel_item=ownership.variant.item,
                           action="Edit",
                           UNKNOWN_COLOR=UNKNOWN_COLOR)


@app.route("/ownership/<int:oid>/delete", methods=["POST"])
def ownership_delete(oid):
    ownership = Ownership.query.get_or_404(oid)
    pid       = ownership.person_id
    db.session.delete(ownership)
    db.session.commit()
    logger.info("Ownership deleted: id %d", oid)
    flash("Entry removed.", "info")
    return redirect(url_for("person_collection", pid=pid))

# ── Wishlist ──────────────────────────────────────────────────────────────────

@app.route("/wishlist")
def wishlist():
    person_id   = request.args.get("person", type=int)
    people_list = Person.query.order_by(Person.name).all()

    wl_q = Ownership.query.filter_by(status="Wishlist")
    if person_id:
        wl_q = wl_q.filter_by(person_id=person_id)
    entries = wl_q.all()

    rows = []
    for entry in entries:
        msrp   = entry.variant.item.msrp
        target = entry.target_price
        hit    = (msrp is not None and target is not None and msrp <= target)
        delta  = (msrp - target) if (msrp is not None and target is not None) else None
        rows.append(dict(
            ownership = entry,
            msrp      = msrp,
            target    = target,
            hit       = hit,
            delta     = delta,
        ))

    # Sort: hits first → closest to target → no price data last
    rows.sort(key=lambda row: (
        0 if row["hit"] else (1 if row["delta"] is not None else 2),
        row["delta"] if row["delta"] is not None else float("inf"),
    ))

    return render_template(
        "wishlist.html",
        rows        = rows,
        people      = people_list,
        person_id   = person_id,
        has_discord = bool(DISCORD_WEBHOOK_URL),
        hit_count   = sum(1 for row in rows if row["hit"]),
    )


@app.route("/wishlist/check", methods=["POST"])
def wishlist_check():
    if not is_admin():
        flash("Admin access required.", "error")
        return redirect(url_for("wishlist"))
    hits = check_wishlist_targets()
    if not hits:
        flash("No wishlist targets met at current MSRP prices.", "info")
        return redirect(url_for("wishlist"))
    if DISCORD_WEBHOOK_URL:
        lines = ["**🎯 Cutco Wishlist — Price Targets Met**"]
        for hit in hits:
            lines.append(
                f"• **{hit['person']}** — {hit['item']} (#{hit['sku']}): "
                f"MSRP ${hit['msrp']:.2f} ≤ target ${hit['target']:.2f} "
                f"(saves ${hit['savings']:.2f})"
            )
        _notify_discord("\n".join(lines))
        flash(f"Sent {len(hits)} price alert(s) to Discord.", "success")
    else:
        flash(
            f"{len(hits)} target(s) met — set DISCORD_WEBHOOK_URL to enable notifications.",
            "info",
        )
    return redirect(url_for("wishlist"))


# ── Sharpening Log ────────────────────────────────────────────────────────────

@app.route("/sharpening")
def sharpening():
    today       = date.today()
    all_entries = (SharpeningLog.query
                   .order_by(SharpeningLog.sharpened_on.desc())
                   .all())

    # Last sharpening date per item
    last_by_item: dict[int, str] = {}
    count_by_item: dict[int, int] = {}
    for entry in all_entries:
        count_by_item[entry.item_id] = count_by_item.get(entry.item_id, 0) + 1
        if entry.item_id not in last_by_item:
            last_by_item[entry.item_id] = entry.sharpened_on

    # Build tracked-items summary
    tracked: list[dict] = []
    for item_id, last_str in last_by_item.items():
        item = Item.query.get(item_id)
        if not item:
            continue
        days_since = (today - date.fromisoformat(last_str)).days
        tracked.append(dict(
            item       = item,
            last_date  = last_str,
            days_since = days_since,
            overdue    = days_since > SHARPEN_THRESHOLD_DAYS,
            event_count= count_by_item[item_id],
        ))

    # Overdue first, then by days desc
    tracked.sort(key=lambda row: (0 if row["overdue"] else 1, -row["days_since"]))

    return render_template(
        "sharpening.html",
        tracked         = tracked,
        recent_entries  = all_entries[:25],
        overdue_count   = sum(1 for row in tracked if row["overdue"]),
        threshold_days  = SHARPEN_THRESHOLD_DAYS,
        today           = today.isoformat(),
        items_list      = Item.query.order_by(Item.name).all(),
        methods         = SHARPEN_METHODS,
        has_discord     = bool(DISCORD_WEBHOOK_URL),
    )


@app.route("/sharpening/add", methods=["POST"])
def sharpening_add():
    item_id      = request.form.get("item_id", type=int)
    sharpened_on = request.form.get("sharpened_on", "").strip()
    method       = request.form.get("method", "Home Sharpener").strip()
    notes        = request.form.get("notes", "").strip() or None

    if not item_id or not sharpened_on:
        flash("Item and date are required.", "error")
        return redirect(url_for("sharpening"))

    if not Item.query.get(item_id):
        flash("Item not found.", "error")
        return redirect(url_for("sharpening"))

    db.session.add(SharpeningLog(
        item_id      = item_id,
        sharpened_on = sharpened_on,
        method       = method,
        notes        = notes,
    ))
    db.session.commit()
    logger.info("Sharpening logged: item %d on %s (%s)", item_id, sharpened_on, method)
    flash("Sharpening event logged.", "success")
    return redirect(url_for("sharpening"))


@app.route("/sharpening/<int:lid>/edit", methods=["GET", "POST"])
def sharpening_edit(lid):
    entry = SharpeningLog.query.get_or_404(lid)
    if request.method == "POST":
        entry.sharpened_on = request.form.get("sharpened_on", entry.sharpened_on).strip()
        entry.method       = request.form.get("method", entry.method).strip()
        entry.notes        = request.form.get("notes", "").strip() or None
        db.session.commit()
        logger.info("Sharpening entry %d updated", lid)
        flash("Event updated.", "success")
        return redirect(url_for("sharpening"))
    return render_template("sharpening_edit.html", entry=entry, methods=SHARPEN_METHODS)


@app.route("/sharpening/<int:lid>/delete", methods=["POST"])
def sharpening_delete(lid):
    entry = SharpeningLog.query.get_or_404(lid)
    db.session.delete(entry)
    db.session.commit()
    logger.info("Sharpening entry %d deleted", lid)
    flash("Event removed.", "info")
    return redirect(url_for("sharpening"))


@app.route("/sharpening/notify", methods=["POST"])
def sharpening_notify():
    if not is_admin():
        flash("Admin access required.", "error")
        return redirect(url_for("sharpening"))

    today       = date.today()
    all_entries = SharpeningLog.query.order_by(SharpeningLog.sharpened_on.desc()).all()
    last_by_item: dict[int, str] = {}
    for entry in all_entries:
        if entry.item_id not in last_by_item:
            last_by_item[entry.item_id] = entry.sharpened_on

    overdue = []
    for item_id, last_str in last_by_item.items():
        days_since = (today - date.fromisoformat(last_str)).days
        if days_since > SHARPEN_THRESHOLD_DAYS:
            item = Item.query.get(item_id)
            if item:
                overdue.append((item, days_since))
    overdue.sort(key=lambda pair: pair[1], reverse=True)

    if not overdue:
        flash(f"No knives overdue for sharpening (threshold: {SHARPEN_THRESHOLD_DAYS} days).", "info")
        return redirect(url_for("sharpening"))

    if DISCORD_WEBHOOK_URL:
        lines = [f"**🔪 Cutco Sharpening Reminder — {len(overdue)} overdue**"]
        for item, days in overdue:
            lines.append(f"• {item.name} — {days} days since last sharpening")
        _notify_discord("\n".join(lines))
        flash(f"Sent reminder for {len(overdue)} overdue knife(s) to Discord.", "success")
    else:
        flash(
            f"{len(overdue)} overdue — set DISCORD_WEBHOOK_URL to enable notifications.",
            "info",
        )
    return redirect(url_for("sharpening"))


# ── Bakeware ──────────────────────────────────────────────────────────────────

@app.route("/bakeware")
def bakeware():
    today       = date.today()
    all_sessions = (BakewareSession.query
                    .order_by(BakewareSession.baked_on.desc())
                    .all())

    # Aggregate per item
    last_by_item:   dict[int, str]   = {}
    count_by_item:  dict[int, int]   = {}
    rating_by_item: dict[int, list]  = {}
    for session in all_sessions:
        iid = session.item_id
        count_by_item[iid] = count_by_item.get(iid, 0) + 1
        if iid not in last_by_item:
            last_by_item[iid] = session.baked_on
        if session.rating is not None:
            rating_by_item.setdefault(iid, []).append(session.rating)

    tracked: list[dict] = []
    for iid, last_str in last_by_item.items():
        item = Item.query.get(iid)
        if not item:
            continue
        days_since = (today - date.fromisoformat(last_str)).days
        ratings    = rating_by_item.get(iid, [])
        tracked.append(dict(
            item       = item,
            last_date  = last_str,
            days_since = days_since,
            stale      = days_since > BAKEWARE_THRESHOLD_DAYS,
            session_count = count_by_item[iid],
            avg_rating = round(sum(ratings) / len(ratings), 1) if ratings else None,
        ))

    # Stale items first, then by days desc
    tracked.sort(key=lambda row: (0 if row["stale"] else 1, -row["days_since"]))

    # Bakeware items from catalog that have never been used
    used_ids = set(last_by_item.keys())
    never_used = (Item.query
                  .filter(Item.category.in_(BAKEWARE_CATEGORIES))
                  .filter(Item.id.notin_(used_ids))
                  .order_by(Item.name)
                  .all()) if BAKEWARE_CATEGORIES else []

    # Item selector: bakeware category items first, then rest
    bakeware_items = (Item.query
                      .filter(Item.category.in_(BAKEWARE_CATEGORIES))
                      .order_by(Item.name).all()) if BAKEWARE_CATEGORIES else []
    other_items    = (Item.query
                      .filter(Item.category.notin_(BAKEWARE_CATEGORIES))
                      .order_by(Item.name).all()) if BAKEWARE_CATEGORIES else Item.query.order_by(Item.name).all()

    return render_template(
        "bakeware.html",
        tracked          = tracked,
        recent_sessions  = all_sessions[:25],
        stale_count      = sum(1 for row in tracked if row["stale"]),
        never_used       = never_used,
        threshold_days   = BAKEWARE_THRESHOLD_DAYS,
        today            = today.isoformat(),
        bakeware_items   = bakeware_items,
        other_items      = other_items,
        has_discord      = bool(DISCORD_WEBHOOK_URL),
    )


@app.route("/bakeware/add", methods=["POST"])
def bakeware_add():
    item_id   = request.form.get("item_id", type=int)
    baked_on  = request.form.get("baked_on", "").strip()
    what_made = request.form.get("what_made", "").strip()
    raw_rating = request.form.get("rating", "").strip()
    notes     = request.form.get("notes", "").strip() or None

    if not item_id or not baked_on or not what_made:
        flash("Item, date, and what you made are required.", "error")
        return redirect(url_for("bakeware"))
    if not Item.query.get(item_id):
        flash("Item not found.", "error")
        return redirect(url_for("bakeware"))

    try:
        rating = int(raw_rating) if raw_rating else None
        if rating is not None and not (1 <= rating <= 5):
            rating = None
    except ValueError:
        rating = None

    db.session.add(BakewareSession(
        item_id  = item_id,
        baked_on = baked_on,
        what_made = what_made,
        rating   = rating,
        notes    = notes,
    ))
    db.session.commit()
    logger.info("Bakeware session logged: item %d on %s — %s", item_id, baked_on, what_made)
    flash("Baking session logged.", "success")
    return redirect(url_for("bakeware"))


@app.route("/bakeware/<int:sid>/edit", methods=["GET", "POST"])
def bakeware_edit(sid):
    session = BakewareSession.query.get_or_404(sid)
    if request.method == "POST":
        session.baked_on  = request.form.get("baked_on", session.baked_on).strip()
        session.what_made = request.form.get("what_made", "").strip() or session.what_made
        session.notes     = request.form.get("notes", "").strip() or None
        raw_rating = request.form.get("rating", "").strip()
        try:
            rating = int(raw_rating) if raw_rating else None
            session.rating = rating if (rating is None or 1 <= rating <= 5) else session.rating
        except ValueError:
            pass
        db.session.commit()
        logger.info("Bakeware session %d updated", sid)
        flash("Session updated.", "success")
        return redirect(url_for("bakeware"))
    return render_template("bakeware_edit.html", session=session)


@app.route("/bakeware/<int:sid>/delete", methods=["POST"])
def bakeware_delete(sid):
    session = BakewareSession.query.get_or_404(sid)
    db.session.delete(session)
    db.session.commit()
    logger.info("Bakeware session %d deleted", sid)
    flash("Session removed.", "info")
    return redirect(url_for("bakeware"))


@app.route("/bakeware/notify", methods=["POST"])
def bakeware_notify():
    if not is_admin():
        flash("Admin access required.", "error")
        return redirect(url_for("bakeware"))

    today        = date.today()
    all_sessions = BakewareSession.query.order_by(BakewareSession.baked_on.desc()).all()
    last_by_item: dict[int, str] = {}
    for session in all_sessions:
        if session.item_id not in last_by_item:
            last_by_item[session.item_id] = session.baked_on

    stale = []
    for iid, last_str in last_by_item.items():
        days_since = (today - date.fromisoformat(last_str)).days
        if days_since > BAKEWARE_THRESHOLD_DAYS:
            item = Item.query.get(iid)
            if item:
                stale.append((item, days_since))
    stale.sort(key=lambda pair: pair[1], reverse=True)

    if not stale:
        flash(f"No bakeware unused for >{BAKEWARE_THRESHOLD_DAYS} days.", "info")
        return redirect(url_for("bakeware"))

    if DISCORD_WEBHOOK_URL:
        lines = [f"**🍰 Bakeware Reminder — {len(stale)} item(s) unused**"]
        for item, days in stale:
            lines.append(f"• {item.name} — {days} days since last use")
        _notify_discord("\n".join(lines))
        flash(f"Sent reminder for {len(stale)} idle bakeware item(s) to Discord.", "success")
    else:
        flash(
            f"{len(stale)} item(s) idle — set DISCORD_WEBHOOK_URL to enable notifications.",
            "info",
        )
    return redirect(url_for("bakeware"))


# ── Stats ─────────────────────────────────────────────────────────────────────

@app.route("/stats")
def stats():
    person_id   = request.args.get("person", type=int)
    people_list = Person.query.order_by(Person.name).all()

    # Owned ownerships, optionally filtered to one collector
    owned_q = (
        db.session.query(Ownership)
        .join(ItemVariant, Ownership.variant_id == ItemVariant.id)
        .join(Item, ItemVariant.item_id == Item.id)
        .filter(Ownership.status == "Owned")
    )
    if person_id:
        owned_q = owned_q.filter(Ownership.person_id == person_id)
    owned = owned_q.all()

    # Deduplicate to unique items (one row per Item even if owned in multiple variants)
    owned_item_map: dict[int, Item] = {}
    for ownership in owned:
        item = ownership.variant.item
        if item.id not in owned_item_map:
            owned_item_map[item.id] = item
    owned_items = list(owned_item_map.values())

    # By category — item count and MSRP total
    cat_counts: dict[str, int]   = {}
    cat_values: dict[str, float] = {}
    cat_catalog: dict[str, int]  = {}  # total catalog items per category

    for item in Item.query.all():
        cat = item.category or "Uncategorized"
        cat_catalog[cat] = cat_catalog.get(cat, 0) + 1

    for item in owned_items:
        cat = item.category or "Uncategorized"
        cat_counts[cat] = cat_counts.get(cat, 0) + 1
        if item.msrp:
            cat_values[cat] = cat_values.get(cat, 0.0) + item.msrp

    # By handle color (from variant)
    color_counts: dict[str, int] = {}
    for ownership in owned:
        color = ownership.variant.color
        if color == UNKNOWN_COLOR:
            color = "Unknown"
        color_counts[color] = color_counts.get(color, 0) + 1

    # By edge type
    edge_counts: dict[str, int] = {}
    for item in owned_items:
        edge = item.edge_type or "Unknown"
        edge_counts[edge] = edge_counts.get(edge, 0) + 1

    # Per-collector summary
    collector_rows = []
    for person in people_list:
        p_owned = Ownership.query.filter_by(person_id=person.id, status="Owned").all()
        p_item_ids = {o.variant.item_id for o in p_owned}
        p_items    = Item.query.filter(Item.id.in_(p_item_ids)).all() if p_item_ids else []
        p_value    = sum(i.msrp for i in p_items if i.msrp)
        collector_rows.append(dict(
            id=person.id, name=person.name,
            count=len(p_item_ids), value=p_value,
        ))
    collector_rows.sort(key=lambda row: row["count"], reverse=True)

    total_value  = sum(i.msrp for i in owned_items if i.msrp)
    priced_count = sum(1 for i in owned_items if i.msrp)
    catalog_total = Item.query.count()

    summary = dict(
        owned_items   = len(owned_items),
        owned_entries = len(owned),
        total_value   = total_value,
        priced_count  = priced_count,
        catalog_total = catalog_total,
        coverage_pct  = round(100 * len(owned_items) / catalog_total, 1) if catalog_total else 0,
    )

    # Sort for charts (descending by count/value)
    cat_data   = sorted(cat_counts.items(),  key=lambda kv: kv[1], reverse=True)
    val_data   = sorted(cat_values.items(),  key=lambda kv: kv[1], reverse=True)
    color_data = sorted(color_counts.items(), key=lambda kv: kv[1], reverse=True)[:15]
    edge_data  = sorted(edge_counts.items(), key=lambda kv: kv[1], reverse=True)

    # Coverage: (owned, total) per category for stacked bar
    cov_cats   = sorted(cat_catalog.keys())
    cov_owned  = [cat_counts.get(cat, 0)   for cat in cov_cats]
    cov_gap    = [cat_catalog.get(cat, 0) - cat_counts.get(cat, 0) for cat in cov_cats]

    return render_template(
        "stats.html",
        people=people_list,
        person_id=person_id,
        summary=summary,
        cat_data=cat_data,
        val_data=val_data,
        color_data=color_data,
        edge_data=edge_data,
        collector_rows=collector_rows,
        cov_cats=cov_cats,
        cov_owned=cov_owned,
        cov_gap=cov_gap,
    )


# ── Views ─────────────────────────────────────────────────────────────────────

@app.route("/views/item/<int:iid>")
def item_owners(iid):
    item = Item.query.get_or_404(iid)
    entries = (Ownership.query
               .join(ItemVariant, Ownership.variant_id == ItemVariant.id)
               .filter(ItemVariant.item_id == iid)
               .order_by(Ownership.status).all())
    owner_ids      = {e.person_id for e in entries}
    people_without = (Person.query
                      .filter(~Person.id.in_(owner_ids))
                      .order_by(Person.name).all())
    return render_template("item_owners.html", item=item,
                           entries=entries, people_without=people_without,
                           UNKNOWN_COLOR=UNKNOWN_COLOR)


@app.route("/views/matrix")
def matrix():
    people_list = Person.query.order_by(Person.name).all()
    items_list  = Item.query.order_by(Item.name).all()

    # Build item_lookup: for each (person, item) pair keep the highest-priority ownership
    item_lookup = {}
    for ownership in Ownership.query.all():
        key     = (ownership.person_id, ownership.variant.item_id)
        current = item_lookup.get(key)
        if current is None or STATUS_RANK.get(ownership.status, 9) < STATUS_RANK.get(current.status, 9):
            item_lookup[key] = ownership

    variant_lookup = {(ownership.person_id, ownership.variant_id): ownership
                      for ownership in Ownership.query.all()}

    variants_by_item = {
        item.id: [variant for variant in item.variants if variant.color != UNKNOWN_COLOR] or item.variants
        for item in items_list
    }

    return render_template("matrix.html",
                           people=people_list,
                           items=items_list,
                           item_lookup=item_lookup,
                           variant_lookup=variant_lookup,
                           variants_by_item=variants_by_item,
                           UNKNOWN_COLOR=UNKNOWN_COLOR)

# ── Export ────────────────────────────────────────────────────────────────────

@app.route("/export/csv")
def export_csv():
    rows = (db.session.query(Ownership, ItemVariant, Item, Person)
            .join(ItemVariant, Ownership.variant_id == ItemVariant.id)
            .join(Item,        ItemVariant.item_id   == Item.id)
            .join(Person,      Ownership.person_id   == Person.id)
            .order_by(Person.name, Item.name, ItemVariant.color).all())

    logger.info("CSV export requested: %d rows", len(rows))
    csv_buffer = io.StringIO()
    writer = csv.writer(csv_buffer)
    writer.writerow(["person", "item_name", "sku", "category", "edge_type",
                     "color", "status", "is_unicorn", "notes"])
    for ownership, variant, item, person in rows:
        writer.writerow([person.name, item.name, item.sku or "", item.category or "",
                         item.edge_type, variant.color, ownership.status,
                         "yes" if (variant.is_unicorn or item.is_unicorn) else "no", ownership.notes or ""])
    csv_buffer.seek(0)
    return Response(csv_buffer.getvalue(), mimetype="text/csv",
                    headers={"Content-Disposition":
                             "attachment; filename=cutco_collection.csv"})

# ── Import ────────────────────────────────────────────────────────────────────

TRUTHY = {"yes", "y", "true", "1", "x"}

# Priority order for resolving duplicate ownership rows in the matrix view.
# Lower number = higher priority (Owned wins over Wishlist, etc.)
STATUS_RANK = {"Owned": 0, "Wishlist": 1, "Traded": 2, "Sold": 3}

# Column mapping from spreadsheet header → internal key
XLSX_COL_MAP = {
    # Core item fields
    "name":                  "name",
    "model #":               "sku",
    "model#":                "sku",
    "color":                 "color",
    "category":              "category",
    "edge":                  "edge_type",
    "unicorn?":              "is_unicorn",
    # Ownership
    "owned?":                "owned_raw",
    # Notes fields (merged into notes column)
    "price":                 "_notes_price",
    "gift box":              "_notes_gift_box",
    "sheath":                "_notes_sheath",
    "quantity purchased":    "_notes_qty",
    "given away":            "_notes_given_away",
}
# Set membership columns — key = lowercase spreadsheet header, value = canonical set name
XLSX_SET_COLS = {s.lower(): s for s in SPREADSHEET_SET_COLUMNS}


def _parse_owned_raw(owned_raw: str, default_person: str | None):
    """
    Parse 'Owned?' cell.
    - Truthy → status=Owned, person=default_person
    - Falsy  → status=Wishlist, person=default_person
    - Text that isn't yes/no → treat as person name, status=Owned
    Returns (status, person_name)
    """
    val = owned_raw.strip()
    if val.lower() in TRUTHY:
        return "Owned", default_person
    if val.lower() in {"no", "n", "false", "0", ""}:
        return "Wishlist", default_person
    # Non-boolean string: treated as assigned person
    return "Owned", val or default_person


def _build_notes(row: dict) -> str | None:
    """Combine spreadsheet auxiliary columns (price, gift box, etc.) into a single notes string.

    Ignores blank or placeholder values (0, none, n/a, -).
    Returns None if all columns are empty.
    """
    parts = []
    for key, label in [
        ("_notes_price",     "Price"),
        ("_notes_gift_box",  "Gift Box"),
        ("_notes_sheath",    "Sheath"),
        ("_notes_qty",       "Qty Purchased"),
        ("_notes_given_away","Given Away"),
    ]:
        value = row.get(key, "").strip()
        if value and value not in ("0", "none", "n/a", "-"):
            parts.append(f"{label}: {value}")
    return "; ".join(parts) or None


@app.route("/import/template")
def import_template():
    csv_buffer = io.StringIO()
    writer = csv.writer(csv_buffer)
    writer.writerow(["name", "sku", "color", "edge_type", "is_unicorn",
                     "person", "status", "category", "notes"])
    writer.writerow(["2-3/4\" Paring Knife", "1720", "Classic Brown", "Double-D",
                     "no", "Anthony", "Owned", "Kitchen Knives", ""])
    writer.writerow(["Super Shears", "2137", "Pearl White", "Straight",
                     "no", "Anthony", "Owned", "Kitchen Knives", ""])
    csv_buffer.seek(0)
    return Response(csv_buffer.getvalue(), mimetype="text/csv",
                    headers={"Content-Disposition":
                             "attachment; filename=cutco_import_template.csv"})


@app.route("/import", methods=["GET", "POST"])
def import_page():
    if request.method == "GET":
        return render_template("import_page.html",
                               people=Person.query.order_by(Person.name).all())

    uploaded_file = request.files.get("csvfile")
    if not uploaded_file or not uploaded_file.filename:
        flash("Please choose a file.", "error")
        return render_template("import_page.html",
                               people=Person.query.order_by(Person.name).all())

    person_override = request.form.get("person_override", "").strip() or None
    ext = uploaded_file.filename.rsplit(".", 1)[-1].lower()
    logger.info("Import file received: %s (person override: %s)", uploaded_file.filename, person_override or "none")

    try:
        if ext == "xlsx":
            wb = openpyxl.load_workbook(io.BytesIO(uploaded_file.stream.read()), data_only=True)
            ws = wb.active
            raw_headers = [str(cell.value).strip() if cell.value is not None else ""
                           for cell in ws[1]]
            norm_rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if all(cell_value is None for cell_value in row):
                    continue
                norm_rows.append({raw_headers[col_idx]: str(cell_value).strip() if cell_value is not None else ""
                                  for col_idx, cell_value in enumerate(row)})
            # Keep original casing for set detection; build internal-key row
            parsed_rows = []
            for row in norm_rows:
                out_row = {}
                set_memberships = []
                for orig_key, val in row.items():
                    normalized_key = orig_key.strip().lower()
                    if normalized_key in XLSX_COL_MAP:
                        out_row[XLSX_COL_MAP[normalized_key]] = val
                    elif normalized_key in XLSX_SET_COLS:
                        if val.strip().lower() in TRUTHY:
                            set_memberships.append(XLSX_SET_COLS[normalized_key])
                    else:
                        # Fallback: snake_case normalisation for CSV-style columns
                        out_row[normalized_key.replace(" ", "_")] = val
                out_row["_sets"] = set_memberships
                parsed_rows.append(out_row)
        else:
            stream = io.StringIO(uploaded_file.stream.read().decode("utf-8-sig"))
            reader = csv.DictReader(stream)
            parsed_rows = []
            for row in reader:
                out_row = {k.strip().lower().replace(" ", "_"): v.strip()
                           for k, v in row.items()}
                out_row["_sets"] = []
                parsed_rows.append(out_row)

    except Exception as exc:
        flash(f"Could not parse file: {exc}", "error")
        return render_template("import_page.html",
                               people=Person.query.order_by(Person.name).all())

    # Apply person override
    if person_override:
        for row in parsed_rows:
            row["owned_raw"] = row.get("owned_raw", "yes")
            row["_person_override"] = person_override

    existing_items   = {item.sku.upper(): item for item in Item.query.filter(Item.sku.isnot(None)).all()}
    existing_names   = {item.name.lower(): item for item in Item.query.all()}
    existing_persons = {person.name.lower(): person for person in Person.query.all()}

    already_in_catalog = []
    new_items_list     = []
    likely_unicorns    = []
    ownership_entries  = []
    conflicts          = []
    errors             = []
    seen_skus          = set()

    for row_num, row in enumerate(parsed_rows, start=2):
        name       = row.get("name", "").strip()
        sku        = (row.get("sku", "") or "").strip().upper() or None
        color      = row.get("color", "").strip() or UNKNOWN_COLOR
        edge_type  = row.get("edge_type", "").strip() or "Unknown"
        is_unicorn = row.get("is_unicorn", "").strip().lower() in TRUTHY
        category   = row.get("category", "").strip() or None
        notes      = _build_notes(row) or row.get("notes", "").strip() or None
        set_names  = row.get("_sets", [])

        # Resolve person + status from 'Owned?' column
        owned_raw = row.get("owned_raw", row.get("status", "yes"))
        status, person_name = _parse_owned_raw(owned_raw, row.get("_person_override") or row.get("person", ""))

        if person_override:
            person_name = person_override

        if not name:
            errors.append({"row": row_num, "reason": "Missing name", "data": row})
            continue

        if status not in STATUS_OPTIONS:
            status = "Owned"

        matched_item = None
        if sku and sku in existing_items:
            matched_item = existing_items[sku]
        elif name.lower() in existing_names:
            matched_item = existing_names[name.lower()]

        # Dedup by (sku, color) so the same item with different handle colors
        # each get their own row in the preview.  The commit creates the item
        # once (first row) and adds variants for subsequent rows.
        dedup_key = (sku or name.lower(), color.lower())

        if matched_item:
            already_in_catalog.append({"item": matched_item, "row": row,
                                       "color": color, "person": person_name,
                                       "status": status, "sets": set_names})
        elif dedup_key not in seen_skus:
            seen_skus.add(dedup_key)
            bucket = likely_unicorns if is_unicorn or not sku else new_items_list
            bucket.append({
                "name": name, "sku": sku, "color": color,
                "edge_type": edge_type, "is_unicorn": is_unicorn,
                "category": category, "notes": notes,
                "person": person_name, "status": status,
                "sets": set_names, "row": row_num,
            })

        if person_name and matched_item:
            person_obj = existing_persons.get(person_name.lower())
            if person_obj:
                variant = next((v for v in matched_item.variants
                                if v.color.lower() == color.lower()), None)
                if variant:
                    existing_o = Ownership.query.filter_by(
                        person_id=person_obj.id, variant_id=variant.id).first()
                    if existing_o:
                        if existing_o.status != status:
                            conflicts.append({
                                "person": person_name,
                                "item": matched_item.name,
                                "color": color,
                                "existing_status": existing_o.status,
                                "import_status": status,
                                "oid": existing_o.id,
                            })
                        continue
            ownership_entries.append({
                "person": person_name,
                "item_name": matched_item.name,
                "item_id":   matched_item.id,
                "color":     color,
                "status":    status,
                "notes":     notes,
                "is_new_person": person_name.lower() not in existing_persons,
            })

    return render_template("import_preview.html",
                           already_in_catalog=already_in_catalog,
                           new_items=new_items_list,
                           likely_unicorns=likely_unicorns,
                           ownership_entries=ownership_entries,
                           conflicts=conflicts,
                           errors=errors,
                           edge_types=EDGE_TYPES,
                           status_options=STATUS_OPTIONS,
                           person_override=person_override)


@app.route("/import/confirm", methods=["POST"])
def import_confirm():
    added_items     = 0
    added_ownership = 0
    added_persons   = 0

    existing_items   = {item.sku.upper(): item for item in Item.query.filter(Item.sku.isnot(None)).all()}
    existing_names   = {item.name.lower(): item for item in Item.query.all()}
    existing_persons = {person.name.lower(): person for person in Person.query.all()}

    item_count = int(request.form.get("item_count", 0))
    for i in range(item_count):
        if request.form.get(f"item_accept_{i}") != "on":
            continue

        name        = request.form.get(f"item_name_{i}", "").strip()
        sku         = request.form.get(f"item_sku_{i}", "").strip().upper() or None
        color       = request.form.get(f"item_color_{i}", "").strip() or UNKNOWN_COLOR
        edge_type   = request.form.get(f"item_edge_{i}", "Unknown")
        is_unicorn  = request.form.get(f"item_unicorn_{i}") == "on"
        category    = request.form.get(f"item_category_{i}", "").strip() or None
        notes       = request.form.get(f"item_notes_{i}", "").strip() or None
        person_name = request.form.get(f"item_person_{i}", "").strip()
        status      = request.form.get(f"item_status_{i}", "Owned")
        set_names   = [sname for sname in request.form.get(f"item_sets_{i}", "").split("|") if sname]

        if not name:
            continue

        item = None
        if sku and sku in existing_items:
            item = existing_items[sku]
        elif name.lower() in existing_names:
            item = existing_names[name.lower()]

        if not item:
            item = Item(name=name, sku=sku, category=category,
                        edge_type=edge_type, is_unicorn=False,
                        in_catalog=bool(sku), notes=notes)
            db.session.add(item)
            db.session.flush()
            ensure_unknown_variant(item)
            if sku:
                existing_items[sku] = item
            existing_names[name.lower()] = item
            added_items += 1

        # Assign set memberships
        for sname in set_names:
            item_set = get_or_create_set(sname)
            if item_set not in item.sets:
                item.sets.append(item_set)

        # Find or create the specific color variant, then apply the unicorn flag
        # to that variant only — not to the item (which would affect all colors).
        target_color = color if (color and color != UNKNOWN_COLOR) else UNKNOWN_COLOR
        variant = next((v for v in item.variants
                        if v.color.lower() == target_color.lower()), None)
        if not variant:
            variant = ItemVariant(item_id=item.id, color=target_color, is_unicorn=is_unicorn)
            db.session.add(variant)
            db.session.flush()
        elif is_unicorn and not variant.is_unicorn:
            variant.is_unicorn = True

        if person_name:
            person = existing_persons.get(person_name.lower())
            if not person:
                person = Person(name=person_name)
                db.session.add(person)
                db.session.flush()
                existing_persons[person_name.lower()] = person
                added_persons += 1
            if not Ownership.query.filter_by(person_id=person.id,
                                             variant_id=variant.id).first():
                db.session.add(Ownership(person_id=person.id,
                                         variant_id=variant.id, status=status))
                added_ownership += 1

    own_count = int(request.form.get("own_count", 0))
    for i in range(own_count):
        if request.form.get(f"own_accept_{i}") != "on":
            continue

        item_id     = int(request.form.get(f"own_item_id_{i}", 0))
        person_name = request.form.get(f"own_person_{i}", "").strip()
        color       = request.form.get(f"own_color_{i}", "").strip() or UNKNOWN_COLOR
        status      = request.form.get(f"own_status_{i}", "Owned")
        notes       = request.form.get(f"own_notes_{i}", "").strip() or None

        item = Item.query.get(item_id)
        if not item or not person_name:
            continue

        person = existing_persons.get(person_name.lower())
        if not person:
            person = Person(name=person_name)
            db.session.add(person)
            db.session.flush()
            existing_persons[person_name.lower()] = person
            added_persons += 1

        variant = next((v for v in item.variants
                        if v.color.lower() == color.lower()), None)
        if not variant:
            variant = ItemVariant(item_id=item.id, color=color)
            db.session.add(variant)
            db.session.flush()

        if not Ownership.query.filter_by(person_id=person.id,
                                          variant_id=variant.id).first():
            db.session.add(Ownership(person_id=person.id,
                                     variant_id=variant.id,
                                     status=status, notes=notes))
            added_ownership += 1

    db.session.commit()
    logger.info("Import: %d items, %d ownership, %d persons",
                added_items, added_ownership, added_persons)

    parts = []
    if added_items:
        parts.append(f"{added_items} item{'s' if added_items != 1 else ''}")
    if added_persons:
        parts.append(f"{added_persons} collector{'s' if added_persons != 1 else ''}")
    if added_ownership:
        parts.append(f"{added_ownership} ownership entr{'ies' if added_ownership != 1 else 'y'}")
    flash("Import complete — added " + (", ".join(parts) if parts else "nothing new") + ".", "success")
    return redirect(url_for("catalog"))

# ── MSRP Diff UI ─────────────────────────────────────────────────────────────

@app.route("/admin/msrp-diff")
def msrp_diff_page():
    if not is_admin():
        flash("Admin access required.", "error")
        return redirect(url_for("index"))
    return render_template("msrp_diff_ui.html", job=_read_msrp_job())


@app.route("/admin/msrp-diff/run", methods=["POST"])
def msrp_diff_run():
    if not is_admin():
        flash("Admin access required.", "error")
        return redirect(url_for("index"))
    job = _read_msrp_job()
    if job["status"] == "running":
        flash("A diff is already running.", "warning")
        return redirect(url_for("msrp_diff_page"))
    update_db = request.form.get("update_db") == "on"
    _write_msrp_job({"status": "running", "progress": [], "results": None,
                     "error": None, "update_db": update_db,
                     "started_at": date.today().isoformat(), "finished_at": None})
    threading.Thread(target=_run_msrp_diff_job, args=(update_db,), daemon=True).start()
    return redirect(url_for("msrp_diff_page"))


@app.route("/admin/msrp-diff/status")
def msrp_diff_status():
    if not is_admin():
        return jsonify(error="Unauthorized"), 403
    return jsonify(_read_msrp_job())


# ── Admin auth ────────────────────────────────────────────────────────────────

@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        if request.form.get("token") == ADMIN_TOKEN:
            resp = redirect(url_for("catalog"))
            resp.set_cookie("admin_token", ADMIN_TOKEN, httponly=True, samesite="Lax")
            logger.info("Admin login successful")
            flash("Admin access granted.", "success")
            return resp
        logger.warning("Admin login failed — wrong token")
        flash("Wrong token.", "error")
    return render_template("admin_login.html")


@app.route("/admin/logout")
def admin_logout():
    logger.info("Admin logged out")
    resp = redirect(url_for("index"))
    resp.delete_cookie("admin_token")
    return resp

# ── API ───────────────────────────────────────────────────────────────────────

@app.route("/api/variants/<int:iid>")
def api_variants(iid):
    item = Item.query.get_or_404(iid)
    return jsonify([{"id": v.id, "color": v.color} for v in item.variants])


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=8080, debug=False)
